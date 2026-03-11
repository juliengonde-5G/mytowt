import csv
import io
import logging
import secrets
from datetime import datetime
from typing import Optional

from fastapi import APIRouter, Request, Depends, Form, HTTPException, Query, UploadFile, File
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse, Response
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from app.templating import templates
from sqlalchemy.ext.asyncio import AsyncSession
from app.utils.activity import log_activity, get_client_ip
from sqlalchemy import select, func
from sqlalchemy.orm import selectinload

logger = logging.getLogger(__name__)

from app.database import get_db
from app.auth import get_current_user
from app.models.user import User
from app.models.vessel import Vessel
from app.models.port import Port
from app.models.leg import Leg
from app.models.finance import OpexParameter, PortConfig
from app.models.emission_parameter import EmissionParameter
from app.models.co2_variable import Co2Variable, CO2_DEFAULTS
from app.models.mrv import MrvParameter, MRV_DEFAULTS

ADMIN_ROLES = {"administrateur", "admin", "data_analyst"}
ADMIN_OR_COMMERCIAL = ADMIN_ROLES | {"commercial"}


async def require_admin(user: User = Depends(get_current_user)):
    """Require admin or data_analyst role for settings access."""
    if user.role not in ADMIN_ROLES:
        raise HTTPException(403, detail="Admin access required")
    return user


async def require_admin_or_commercial(user: User = Depends(get_current_user)):
    """Require admin, data_analyst or commercial role."""
    if user.role not in ADMIN_OR_COMMERCIAL:
        raise HTTPException(403, detail="Access denied")
    return user

router = APIRouter(prefix="/admin", tags=["admin"])

# Separate router for routes accessible to all authenticated users (my-account)
# or admin+commercial (settings)
account_router = APIRouter(prefix="/admin", tags=["admin"])

USER_ROLES = [
    {"value": "administrateur", "label": "Administrateur"},
    {"value": "operation", "label": "Opération"},
    {"value": "armement", "label": "Armement"},
    {"value": "technique", "label": "Technique"},
    {"value": "data_analyst", "label": "Data Analyst"},
    {"value": "marins", "label": "Marins"},
    {"value": "commercial", "label": "Commercial"},
    {"value": "manager_maritime", "label": "Manager Maritime"},
]




def pf(val, default=0):
    if val is None or (isinstance(val, str) and val.strip() == ""):
        return default
    try:
        if isinstance(val, str):
            val = val.replace(" ", "").replace("\u00a0", "")
            if "." in val and "," in val:
                val = val.replace(".", "").replace(",", ".")
            elif "," in val:
                val = val.replace(",", ".")
        return float(val)
    except (ValueError, TypeError):
        return default


# ─── SETTINGS HOME ───────────────────────────────────────────
@account_router.get("/settings", response_class=HTMLResponse)
async def settings_home(
    request: Request,
    user: User = Depends(require_admin_or_commercial),
    db: AsyncSession = Depends(get_db),
):
    # Users
    users_result = await db.execute(select(User).order_by(User.id))
    users = users_result.scalars().all()

    # Vessels
    vessels_result = await db.execute(select(Vessel).order_by(Vessel.code))
    vessels = vessels_result.scalars().all()

    # Port configs
    ports_result = await db.execute(
        select(PortConfig).options(selectinload(PortConfig.port)).order_by(PortConfig.port_locode)
    )
    port_configs = ports_result.scalars().all()

    # OPEX
    opex_result = await db.execute(select(OpexParameter).order_by(OpexParameter.parameter_name))
    opex_params = opex_result.scalars().all()

    # Emission params
    emission_result = await db.execute(select(EmissionParameter).order_by(EmissionParameter.parameter_name))
    emission_params = emission_result.scalars().all()

    # Locked legs count
    locked_result = await db.execute(
        select(func.count(Leg.id)).where(Leg.status == "completed")
    )
    locked_count = locked_result.scalar() or 0

    # Cabin pricing grid
    from app.models.passenger import CabinPriceGrid, CABIN_TYPE_LABELS
    pricing_result = await db.execute(
        select(CabinPriceGrid).order_by(CabinPriceGrid.origin_locode, CabinPriceGrid.destination_locode, CabinPriceGrid.cabin_type)
    )
    pricing_grid = pricing_result.scalars().all()

    # Existing routes (unique origin→dest pairs from legs)
    from sqlalchemy.orm import selectinload as sl
    routes_result = await db.execute(
        select(
            Leg.departure_port_locode, Leg.arrival_port_locode
        ).distinct()
    )
    route_pairs = routes_result.all()

    # Collect unique port locodes used in routes
    route_locodes = set()
    for dep, arr in route_pairs:
        route_locodes.add(dep)
        route_locodes.add(arr)

    # Load port names for those locodes
    route_ports = {}
    if route_locodes:
        ports_result = await db.execute(
            select(Port).where(Port.locode.in_(route_locodes)).order_by(Port.name)
        )
        for p in ports_result.scalars().all():
            route_ports[p.locode] = p.name

    # Build structured routes list for template
    existing_routes = sorted(
        [{"dep": dep, "arr": arr, "dep_name": route_ports.get(dep, dep), "arr_name": route_ports.get(arr, arr)}
         for dep, arr in route_pairs],
        key=lambda r: (r["dep_name"], r["arr_name"])
    )

    # Insurance contracts
    from app.models.finance import InsuranceContract
    ins_result = await db.execute(select(InsuranceContract).order_by(InsuranceContract.guarantee_type))
    insurance_contracts = ins_result.scalars().all()

    # CO2 variables
    co2_result = await db.execute(
        select(Co2Variable).where(Co2Variable.is_current == True).order_by(Co2Variable.variable_name)
    )
    co2_vars = co2_result.scalars().all()
    co2_dict = {v.variable_name: v for v in co2_vars}

    # CO2 history (TOWT EF only)
    co2_hist_result = await db.execute(
        select(Co2Variable).where(Co2Variable.variable_name == "towt_co2_ef")
        .order_by(Co2Variable.effective_date.desc())
    )
    co2_history = co2_hist_result.scalars().all()

    # MRV params
    mrv_result = await db.execute(select(MrvParameter).order_by(MrvParameter.parameter_name))
    mrv_params = mrv_result.scalars().all()

    # Pipedrive token
    pd_result = await db.execute(
        select(OpexParameter).where(OpexParameter.parameter_name == "pipedrive_api_token")
    )
    pd_param = pd_result.scalar_one_or_none()
    pipedrive_token = pd_param.description if pd_param else ""  # stored in description (text field)

    return templates.TemplateResponse("admin/settings.html", {
        "request": request, "user": user,
        "users": users, "vessels": vessels,
        "port_configs": port_configs,
        "opex_params": opex_params,
        "emission_params": emission_params,
        "locked_count": locked_count,
        "pricing_grid": pricing_grid,
        "existing_routes": existing_routes,
        "route_ports": route_ports,
        "cabin_type_labels": CABIN_TYPE_LABELS,
        "insurance_contracts": insurance_contracts,
        "co2_vars": co2_vars, "co2_dict": co2_dict, "co2_history": co2_history,
        "co2_defaults": CO2_DEFAULTS,
        "mrv_params": mrv_params, "mrv_defaults": MRV_DEFAULTS,
        "pipedrive_token": pipedrive_token,
        "active_module": "settings",
    })


# ─── USERS CRUD ──────────────────────────────────────────────
@router.get("/users", response_class=HTMLResponse)
async def users_list(
    request: Request, user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(select(User).order_by(User.id))
    users = result.scalars().all()
    return templates.TemplateResponse("admin/users.html", {
        "request": request, "user": user, "users": users,
        "active_module": "settings",
    })


@router.get("/users/create", response_class=HTMLResponse)
async def user_create_form(request: Request, user: User = Depends(get_current_user)):
    return templates.TemplateResponse("admin/user_form.html", {
        "request": request, "user": user, "edit_user": None, "error": None, "roles": USER_ROLES,
    })


@router.post("/users/create", response_class=HTMLResponse)
async def user_create_submit(
    request: Request,
    username: str = Form(...), password: str = Form(...),
    full_name: str = Form(""), role: str = Form("operation"),
    email: str = Form(""),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    from app.auth import hash_password
    # Input length validation
    username = username[:50]
    password = password[:128]
    full_name = full_name[:100]
    email = email[:200]
    if len(username) < 3:
        return templates.TemplateResponse("admin/user_form.html", {
            "request": request, "user": user, "edit_user": None,
            "error": "Le nom d'utilisateur doit contenir au moins 3 caractères.", "roles": USER_ROLES,
        })
    if len(password) < 8:
        return templates.TemplateResponse("admin/user_form.html", {
            "request": request, "user": user, "edit_user": None,
            "error": "Le mot de passe doit contenir au moins 8 caractères.", "roles": USER_ROLES,
        })
    if not email:
        email = f"{username}@towt.eu"
    existing = await db.execute(select(User).where(User.username == username))
    if existing.scalar_one_or_none():
        return templates.TemplateResponse("admin/user_form.html", {
            "request": request, "user": user, "edit_user": None,
            "error": "Ce nom d'utilisateur existe déjà.", "roles": USER_ROLES,
        })
    pwd_hash = hash_password(password)
    new_user = User(username=username, email=email, hashed_password=pwd_hash, full_name=full_name, role=role)
    db.add(new_user)
    await log_activity(db, user=user, action="create", module="admin",
                       entity_type="user", entity_id=new_user.id,
                       entity_label=new_user.full_name, detail=f"role: {new_user.role}",
                       ip_address=get_client_ip(request))
    await db.flush()
    return RedirectResponse(url="/admin/users", status_code=303)


@router.get("/users/{uid}/edit", response_class=HTMLResponse)
async def user_edit_form(
    uid: int, request: Request,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(select(User).where(User.id == uid))
    edit_user = result.scalar_one_or_none()
    if not edit_user:
        raise HTTPException(status_code=404)
    return templates.TemplateResponse("admin/user_form.html", {
        "request": request, "user": user, "edit_user": edit_user, "error": None, "roles": USER_ROLES,
    })


@router.post("/users/{uid}/edit", response_class=HTMLResponse)
async def user_edit_submit(
    uid: int, request: Request,
    full_name: str = Form(""), role: str = Form("viewer"),
    password: Optional[str] = Form(None),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    from app.auth import hash_password as _hp
    result = await db.execute(select(User).where(User.id == uid))
    edit_user = result.scalar_one_or_none()
    if not edit_user:
        raise HTTPException(status_code=404)
    edit_user.full_name = full_name[:100]
    edit_user.role = role
    if password and password.strip():
        edit_user.hashed_password = _hp(password[:128])
    await db.flush()
    await log_activity(db, user=user, action="update", module="admin",
                       entity_type="user", entity_id=edit_user.id,
                       entity_label=edit_user.full_name, detail=f"role: {role}",
                       ip_address=get_client_ip(request))
    return RedirectResponse(url="/admin/users", status_code=303)


@router.delete("/users/{uid}", response_class=HTMLResponse)
async def user_delete(
    uid: int, request: Request,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(select(User).where(User.id == uid))
    u = result.scalar_one_or_none()
    if u and u.id != user.id:
        await log_activity(db, user=user, action="delete", module="admin",
                           entity_type="user", entity_id=uid, entity_label=u.full_name,
                           ip_address=get_client_ip(request))
        await db.delete(u)
        await db.flush()
    if request.headers.get("HX-Request"):
        return HTMLResponse(content="", headers={"HX-Redirect": "/admin/users"})
    return RedirectResponse(url="/admin/users", status_code=303)


# ─── USER IMPORT (EXCEL) ────────────────────────────────────
VALID_ROLES = {r["value"] for r in USER_ROLES}

@router.get("/users/import/template")
async def users_import_template(user: User = Depends(require_admin)):
    """Download an Excel template for user import."""
    wb = openpyxl.Workbook()

    # ── Sheet 1: Template ──
    ws = wb.active
    ws.title = "Utilisateurs"
    headers = ["username", "full_name", "email", "role", "password", "language", "is_active"]
    header_font = Font(name="Poppins", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="095561", end_color="095561", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Example rows
    examples = [
        ("jdupont", "Jean Dupont", "jean.dupont@towt.eu", "operation", "MotDePasse123", "fr", "oui"),
        ("mmartin", "Marie Martin", "marie.martin@towt.eu", "commercial", "MotDePasse456", "en", "oui"),
    ]
    example_font = Font(name="Poppins", color="888888", italic=True, size=10)
    for r, row_data in enumerate(examples, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = example_font
            cell.border = thin_border

    # Column widths
    widths = [18, 25, 30, 22, 20, 12, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # ── Sheet 2: Référence rôles ──
    ws2 = wb.create_sheet("Référence rôles")
    ws2.cell(row=1, column=1, value="Valeur (role)").font = Font(bold=True)
    ws2.cell(row=1, column=2, value="Description").font = Font(bold=True)
    for i, r in enumerate(USER_ROLES, 2):
        ws2.cell(row=i, column=1, value=r["value"])
        ws2.cell(row=i, column=2, value=r["label"])
    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 30

    # ── Sheet 3: Instructions ──
    ws3 = wb.create_sheet("Instructions")
    instructions = [
        "Instructions d'import des utilisateurs",
        "",
        "Colonnes obligatoires : username, full_name, password",
        "Colonnes optionnelles : email, role, language, is_active",
        "",
        "username : identifiant unique (3-50 car.), pas d'espaces",
        "full_name : nom complet (max 100 car.)",
        "email : si vide, sera généré comme username@towt.eu",
        "role : voir onglet 'Référence rôles' (défaut: operation)",
        "password : min 8 caractères",
        "language : fr, en, es, pt-br, vi (défaut: fr)",
        "is_active : oui/non ou true/false (défaut: oui)",
        "",
        "Les utilisateurs existants (même username) seront ignorés.",
        "Les lignes avec erreurs seront signalées dans le rapport.",
    ]
    for i, line in enumerate(instructions, 1):
        cell = ws3.cell(row=i, column=1, value=line)
        if i == 1:
            cell.font = Font(bold=True, size=13)
        else:
            cell.font = Font(size=11)
    ws3.column_dimensions["A"].width = 60

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=import_utilisateurs_template.xlsx"},
    )


@router.post("/users/import", response_class=HTMLResponse)
async def users_import_excel(
    request: Request,
    file: UploadFile = File(...),
    user: User = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    """Import users from an Excel file."""
    from app.auth import hash_password

    raw = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except Exception:
        msg = "Fichier Excel invalide. Utilisez le template fourni."
        return RedirectResponse(url=f"/admin/users?import_error={msg.replace(' ', '+')}", status_code=303)

    ws = wb.active
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    if len(rows) < 2:
        return RedirectResponse(url="/admin/users?import_error=Fichier+vide+ou+sans+données", status_code=303)

    # Map header names to column indices
    raw_headers = [str(h).strip().lower() if h else "" for h in rows[0]]
    col_map = {}
    for i, h in enumerate(raw_headers):
        col_map[h] = i

    required = {"username", "full_name", "password"}
    missing = required - set(col_map.keys())
    if missing:
        msg = f"Colonnes manquantes : {', '.join(missing)}"
        return RedirectResponse(url=f"/admin/users?import_error={msg.replace(' ', '+')}", status_code=303)

    def cell_val(row, col_name, default=""):
        idx = col_map.get(col_name)
        if idx is None or idx >= len(row) or row[idx] is None:
            return default
        return str(row[idx]).strip()

    # Preload existing usernames
    result = await db.execute(select(User.username))
    existing_usernames = {u.lower() for u in result.scalars().all()}

    created = 0
    skipped = 0
    errors = []

    for line_no, row in enumerate(rows[1:], 2):
        # Skip empty rows
        if not any(row):
            continue

        username = cell_val(row, "username")
        full_name = cell_val(row, "full_name")
        password = cell_val(row, "password")
        email = cell_val(row, "email")
        role = cell_val(row, "role", "operation").lower()
        language = cell_val(row, "language", "fr").lower()
        is_active_str = cell_val(row, "is_active", "oui").lower()

        # Validate
        if not username or len(username) < 3:
            errors.append(f"Ligne {line_no}: username manquant ou trop court (<3 car.)")
            continue
        username = username[:50]

        if not full_name:
            errors.append(f"Ligne {line_no}: full_name manquant")
            continue
        full_name = full_name[:100]

        if not password or len(password) < 8:
            errors.append(f"Ligne {line_no}: password manquant ou trop court (<8 car.)")
            continue
        password = password[:128]

        if username.lower() in existing_usernames:
            skipped += 1
            continue

        if role not in VALID_ROLES:
            errors.append(f"Ligne {line_no}: rôle '{role}' invalide")
            continue

        if language not in ("fr", "en", "es", "pt-br", "vi"):
            language = "fr"

        if not email:
            email = f"{username}@towt.eu"
        email = email[:255]

        is_active = is_active_str not in ("non", "false", "0", "no", "inactif")

        new_user = User(
            username=username,
            email=email,
            hashed_password=hash_password(password),
            full_name=full_name,
            role=role,
            language=language,
            is_active=is_active,
        )
        db.add(new_user)
        existing_usernames.add(username.lower())
        created += 1

    await db.flush()
    await log_activity(db, user=user, action="import", module="admin",
                       entity_type="users_excel", entity_id=None,
                       entity_label=f"{file.filename}: {created} créés, {skipped} ignorés",
                       ip_address=get_client_ip(request))

    msg = f"Import terminé : {created} utilisateurs créés, {skipped} existants ignorés"
    if errors:
        msg += f", {len(errors)} erreurs"
    param = "import_success" if created > 0 or (not errors) else "import_error"
    url = f"/admin/users?{param}={msg.replace(' ', '+')}"
    if errors:
        detail = " | ".join(errors[:10])
        if len(errors) > 10:
            detail += f" ... et {len(errors)-10} autres"
        url += f"&import_detail={detail.replace(' ', '+')}"

    if request.headers.get("HX-Request"):
        return HTMLResponse(content="", headers={"HX-Redirect": url})
    return RedirectResponse(url=url, status_code=303)


# ─── VESSELS ─────────────────────────────────────────────────
@router.get("/vessels/create", response_class=HTMLResponse)
async def vessel_create_form(
    request: Request,
    user: User = Depends(get_current_user),
):
    return templates.TemplateResponse("admin/vessel_form.html", {
        "request": request, "user": user, "vessel": None,
    })


@router.post("/vessels/create", response_class=HTMLResponse)
async def vessel_create_submit(
    request: Request,
    code: int = Form(...),
    name: str = Form(...),
    imo_number: str = Form(""),
    flag: str = Form(""),
    dwt: str = Form(""),
    capacity_palettes: str = Form("0"),
    default_speed: str = Form("8"),
    default_elongation: str = Form("1.25"),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    # Check unique code
    existing = await db.execute(select(Vessel).where(Vessel.code == code))
    if existing.scalar_one_or_none():
        return templates.TemplateResponse("admin/vessel_form.html", {
            "request": request, "user": user, "vessel": None,
            "error": "Ce code navire existe déjà.",
        })
    vessel = Vessel(
        code=code, name=name.strip(),
        imo_number=imo_number.strip() or None,
        flag=flag.strip() or None,
        dwt=pf(dwt, 0) if dwt else None,
        capacity_palettes=int(pf(capacity_palettes, 0)),
        default_speed=pf(default_speed, 8),
        default_elongation=pf(default_elongation, 1.25),
    )
    db.add(vessel)
    await db.flush()
    await log_activity(db, user=user, action="create", module="admin",
                       entity_type="vessel", entity_id=vessel.id,
                       entity_label=vessel.name, ip_address=get_client_ip(request))
    if request.headers.get("HX-Request"):
        return HTMLResponse(content="", headers={"HX-Redirect": "/admin/settings#vessels"})
    return RedirectResponse(url="/admin/settings#vessels", status_code=303)


@router.get("/vessels/{vid}/edit", response_class=HTMLResponse)
async def vessel_edit_form(
    vid: int, request: Request,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(select(Vessel).where(Vessel.id == vid))
    vessel = result.scalar_one_or_none()
    if not vessel:
        raise HTTPException(status_code=404)
    return templates.TemplateResponse("admin/vessel_form.html", {
        "request": request, "user": user, "vessel": vessel,
    })


@router.post("/vessels/{vid}/edit", response_class=HTMLResponse)
async def vessel_edit_submit(
    vid: int, request: Request,
    name: str = Form(...),
    capacity_palettes: str = Form("0"),
    default_speed: str = Form("8"),
    default_elongation: str = Form("1.25"),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(select(Vessel).where(Vessel.id == vid))
    vessel = result.scalar_one_or_none()
    if not vessel:
        raise HTTPException(status_code=404)
    vessel.name = name
    vessel.capacity_palettes = int(pf(capacity_palettes, 0))
    vessel.default_speed = pf(default_speed, 8)
    vessel.default_elongation = pf(default_elongation, 1.25)
    await db.flush()
    await log_activity(db, user=user, action="update", module="admin",
                       entity_type="vessel", entity_id=vessel.id,
                       entity_label=vessel.name, ip_address=get_client_ip(request))
    if request.headers.get("HX-Request"):
        return HTMLResponse(content="", headers={"HX-Redirect": "/admin/settings"})
    return RedirectResponse(url="/admin/settings", status_code=303)


# ─── OPEX PARAMETER ──────────────────────────────────────────
@router.post("/opex/update", response_class=HTMLResponse)
async def opex_update(
    request: Request,
    opex_daily_rate: str = Form("11600"),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(
        select(OpexParameter).where(OpexParameter.parameter_name == "opex_daily_rate")
    )
    param = result.scalar_one_or_none()
    if param:
        param.parameter_value = pf(opex_daily_rate, 11600)
    else:
        param = OpexParameter(
            parameter_name="opex_daily_rate", parameter_value=pf(opex_daily_rate, 11600),
            unit="EUR/jour", category="global", description="Coût journalier en mer",
        )
        db.add(param)
    await db.flush()
    await log_activity(db, user=user, action="update", module="admin",
                       entity_type="opex", detail=f"opex_daily_rate={opex_daily_rate}",
                       ip_address=get_client_ip(request))
    if request.headers.get("HX-Request"):
        return HTMLResponse(content="", headers={"HX-Redirect": "/admin/settings"})
    return RedirectResponse(url="/admin/settings", status_code=303)


# ─── LOCK/UNLOCK LEGS ────────────────────────────────────────
@router.post("/legs/lock", response_class=HTMLResponse)
async def lock_completed_legs(
    request: Request,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Lock all legs where ATD is set (voyage completed)."""
    result = await db.execute(select(Leg).where(Leg.atd != None, Leg.status != "completed"))
    legs = result.scalars().all()
    count = 0
    for leg in legs:
        leg.status = "completed"
        count += 1
    await db.flush()
    if request.headers.get("HX-Request"):
        return HTMLResponse(content="", headers={"HX-Redirect": f"/admin/settings?locked={count}"})
    return RedirectResponse(url="/admin/settings", status_code=303)


@router.post("/legs/{lid}/unlock", response_class=HTMLResponse)
async def unlock_leg(
    lid: int, request: Request,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(select(Leg).where(Leg.id == lid))
    leg = result.scalar_one_or_none()
    if leg:
        leg.status = "planned"
        await db.flush()
    if request.headers.get("HX-Request"):
        return HTMLResponse(content="", headers={"HX-Redirect": "/admin/settings"})
    return RedirectResponse(url="/admin/settings", status_code=303)


# ─── DATABASE MANAGEMENT ─────────────────────────────────────

# Helper: export a table to CSV
def _table_to_csv(rows, columns):
    output = io.StringIO()
    writer = csv.writer(output, delimiter=";")
    writer.writerow(columns)
    for row in rows:
        writer.writerow([getattr(row, c, '') for c in columns])
    return output.getvalue()

def _date_fmt(val):
    if val is None: return ''
    try: return val.strftime('%d/%m/%Y %H:%M')
    except: return str(val)

# Table definitions for export/purge
TABLE_DEFS = {
    "legs": {"label": "Legs", "tables": ["legs"]},
    "orders": {"label": "Commandes", "tables": ["orders", "order_assignments"]},
    "passengers": {"label": "Passagers", "tables": ["passenger_bookings", "passengers", "passenger_payments", "passenger_documents", "preboarding_forms", "passenger_audit_logs"]},
    "cargo": {"label": "Cargo", "tables": ["packing_lists", "packing_list_batches", "packing_list_audit"]},
    "finance": {"label": "Finance", "tables": ["leg_finances"]},
    "claims": {"label": "Claims", "tables": ["claims", "claim_documents", "claim_timeline"]},
    "crew": {"label": "Équipage", "tables": ["crew_members", "crew_assignments", "crew_tickets"]},
    "crew_assignments": {"label": "Affectations", "tables": ["crew_assignments"]},
    "escale": {"label": "Escale", "tables": ["escale_operations", "docker_shifts"]},
    "sof": {"label": "SOF", "tables": ["sof_events"]},
    "onboard": {"label": "On Board", "tables": ["onboard_notifications", "onboard_attachments", "cargo_documents", "cargo_document_attachments", "eta_shifts"]},
    "messages": {"label": "Messages", "tables": ["portal_messages"]},
    "notifications": {"label": "Notifications", "tables": ["notifications"]},
    "activity": {"label": "Journal d'activité", "tables": ["activity_logs"]},
    "clients": {"label": "Clients", "tables": ["clients"]},
    "rates": {"label": "Grilles tarifaires", "tables": ["rate_grids", "rate_grid_lines", "rate_offers"]},
    "kpi": {"label": "KPI", "tables": ["leg_kpis"]},
    "config": {"label": "Config", "tables": ["vessels", "ports", "port_configs", "opex_parameters", "emission_parameters", "cabin_price_grid", "insurance_contracts", "co2_variables", "mrv_parameters", "mrv_events", "planning_shares", "vessel_positions"]},
    "access_logs": {"label": "Logs d'accès", "tables": ["portal_access_logs"]},
}

# Whitelist of all allowed table names for SQL queries (prevents SQL injection)
ALLOWED_TABLES = set()
for _def in TABLE_DEFS.values():
    ALLOWED_TABLES.update(_def["tables"])


@router.get("/export/global")
async def export_global(
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    import zipfile
    from sqlalchemy import text
    zip_buffer = io.BytesIO()
    now_str = datetime.now().strftime("%Y%m%d_%H%M")

    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
        # Get all table names from the database
        result = await db.execute(text(
            "SELECT tablename FROM pg_tables WHERE schemaname = 'public' ORDER BY tablename"
        ))
        tables = [r[0] for r in result.fetchall()]

        for table_name in tables:
            if table_name not in ALLOWED_TABLES:
                continue
            try:
                cols_result = await db.execute(
                    text("SELECT column_name FROM information_schema.columns WHERE table_name = :tname ORDER BY ordinal_position").bindparams(tname=table_name)
                )
                columns = [r[0] for r in cols_result.fetchall()]
                rows_result = await db.execute(text(f'SELECT * FROM "{table_name}" LIMIT 50000'))
                rows = rows_result.fetchall()

                output = io.StringIO()
                writer = csv.writer(output, delimiter=";")
                writer.writerow(columns)
                for row in rows:
                    writer.writerow([_date_fmt(v) if hasattr(v, 'strftime') else ('' if v is None else str(v)) for v in row])

                zf.writestr(f"{table_name}.csv", output.getvalue())
            except Exception as e:
                logger.warning(f"Export failed for table {table_name}: {e}")
                continue

    zip_buffer.seek(0)
    return StreamingResponse(
        iter([zip_buffer.getvalue()]),
        media_type="application/zip",
        headers={"Content-Disposition": f"attachment; filename=towt_export_complet_{now_str}.zip"},
    )


@router.get("/export/selective")
async def export_selective(
    request: Request,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    import zipfile
    from sqlalchemy import text
    tables_param = request.query_params.getlist("tables")
    if not tables_param:
        return RedirectResponse(url="/admin/settings#database", status_code=303)

    zip_buffer = io.BytesIO()
    now_str = datetime.now().strftime("%Y%m%d_%H%M")
    selected_tables = set()
    for key in tables_param:
        if key in TABLE_DEFS:
            selected_tables.update(TABLE_DEFS[key]["tables"])

    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
        for table_name in sorted(selected_tables):
            if table_name not in ALLOWED_TABLES:
                continue
            try:
                cols_result = await db.execute(
                    text("SELECT column_name FROM information_schema.columns WHERE table_name = :tname ORDER BY ordinal_position").bindparams(tname=table_name)
                )
                columns = [r[0] for r in cols_result.fetchall()]
                if not columns:
                    continue
                rows_result = await db.execute(text(f'SELECT * FROM "{table_name}" LIMIT 50000'))
                rows = rows_result.fetchall()

                output = io.StringIO()
                writer = csv.writer(output, delimiter=";")
                writer.writerow(columns)
                for row in rows:
                    writer.writerow([_date_fmt(v) if hasattr(v, 'strftime') else ('' if v is None else str(v)) for v in row])

                zf.writestr(f"{table_name}.csv", output.getvalue())
            except Exception as e:
                logger.warning(f"Selective export failed for table {table_name}: {e}")
                continue

    zip_buffer.seek(0)
    return StreamingResponse(
        iter([zip_buffer.getvalue()]),
        media_type="application/zip",
        headers={"Content-Disposition": f"attachment; filename=towt_export_selection_{now_str}.zip"},
    )


@router.get("/export/files")
async def export_files(
    user: User = Depends(get_current_user),
):
    import zipfile, os
    zip_buffer = io.BytesIO()
    now_str = datetime.now().strftime("%Y%m%d_%H%M")

    upload_dirs = [
        ("/app/uploads/passenger_docs", "passenger_docs"),
        ("/app/data/claims", "claims"),
        ("app/static/uploads/orders", "orders"),
        ("/app/uploads/crew_tickets", "crew_tickets"),
        ("/app/uploads/onboard_attachments", "onboard_attachments"),
        ("/app/uploads/cargo_doc_attachments", "cargo_doc_attachments"),
    ]

    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
        for dir_path, prefix in upload_dirs:
            if os.path.exists(dir_path):
                for root, dirs, files in os.walk(dir_path):
                    for fname in files:
                        full_path = os.path.join(root, fname)
                        arc_name = os.path.join(prefix, os.path.relpath(full_path, dir_path))
                        try:
                            zf.write(full_path, arc_name)
                        except Exception as e:
                            logger.warning(f"File export failed for {full_path}: {e}")
                            continue

    zip_buffer.seek(0)
    return StreamingResponse(
        iter([zip_buffer.getvalue()]),
        media_type="application/zip",
        headers={"Content-Disposition": f"attachment; filename=towt_fichiers_{now_str}.zip"},
    )


@router.post("/database/purge-selective", response_class=HTMLResponse)
async def purge_selective(
    request: Request,
    confirm_text: str = Form(""),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    if confirm_text.strip() != "SUPPRIMER":
        return RedirectResponse(url="/admin/settings#database", status_code=303)
    form = await request.form()
    tables_param = form.getlist("tables")
    if not tables_param:
        return RedirectResponse(url="/admin/settings#database", status_code=303)

    from sqlalchemy import text

    # Order matters for FK constraints — delete children first
    purge_order = [
        "notifications", "messages", "activity", "access_logs", "sof",
        "onboard", "escale", "claims",
        "cargo", "passengers", "finance", "rates", "clients",
        "crew_assignments", "orders", "kpi", "legs",
    ]

    table_sql = {
        "notifications": ["DELETE FROM notifications"],
        "messages": ["DELETE FROM portal_messages"],
        "activity": ["DELETE FROM activity_logs"],
        "access_logs": ["DELETE FROM portal_access_logs"],
        "sof": ["DELETE FROM sof_events WHERE event_type NOT IN ('CLAIM_DECLARED','CLAIM_UPDATED')"],
        "onboard": [
            "DELETE FROM cargo_document_attachments", "DELETE FROM cargo_documents",
            "DELETE FROM onboard_attachments", "DELETE FROM onboard_notifications",
            "DELETE FROM eta_shifts",
        ],
        "escale": ["DELETE FROM escale_operations", "DELETE FROM docker_shifts"],
        "claims": ["DELETE FROM claim_timeline", "DELETE FROM claim_documents", "DELETE FROM claims"],
        "cargo": ["DELETE FROM packing_list_audit", "DELETE FROM packing_list_batches", "DELETE FROM packing_lists"],
        "passengers": [
            "DELETE FROM passenger_audit_logs", "DELETE FROM preboarding_forms",
            "DELETE FROM passenger_documents", "DELETE FROM passenger_payments",
            "DELETE FROM passengers", "DELETE FROM passenger_bookings",
        ],
        "orders": ["DELETE FROM order_assignments", "DELETE FROM orders"],
        "finance": ["DELETE FROM leg_finances"],
        "rates": ["DELETE FROM rate_offers", "DELETE FROM rate_grid_lines", "DELETE FROM rate_grids"],
        "clients": ["DELETE FROM clients"],
        "crew_assignments": ["DELETE FROM crew_tickets", "DELETE FROM crew_assignments"],
        "kpi": ["DELETE FROM leg_kpis"],
        "legs": [
            "DELETE FROM sof_events", "DELETE FROM escale_operations",
            "DELETE FROM onboard_notifications", "DELETE FROM cargo_documents",
            "DELETE FROM cargo_document_attachments",
            "DELETE FROM docker_shifts", "DELETE FROM leg_finances", "DELETE FROM leg_kpis",
            "DELETE FROM eta_shifts", "DELETE FROM onboard_attachments",
            "DELETE FROM crew_tickets",
            "DELETE FROM legs",
        ],
    }

    for key in purge_order:
        if key in tables_param and key in table_sql:
            for sql in table_sql[key]:
                try:
                    await db.execute(text(sql))
                except Exception as e:
                    logger.warning(f"Purge SQL failed ({sql[:50]}): {e}")
    await log_activity(db, user=user, action="purge", module="admin",
                       detail=f"Purge sélective: {', '.join(tables_param)}",
                       ip_address=get_client_ip(request))
    await db.flush()
    return RedirectResponse(url="/admin/settings#database", status_code=303)


@router.post("/database/reset", response_class=HTMLResponse)
async def reset_database(
    request: Request,
    confirm_text: str = Form(""),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    form = await request.form()
    if confirm_text.strip() != "REINITIALISER" or not form.get("confirm_checkbox"):
        return RedirectResponse(url="/admin/settings#database", status_code=303)

    from sqlalchemy import text

    # Delete everything in dependency order, keep users + config
    purge_sql = [
        "DELETE FROM notifications",
        "DELETE FROM portal_messages",
        "DELETE FROM portal_access_logs",
        "DELETE FROM activity_logs",
        "DELETE FROM claim_timeline",
        "DELETE FROM claim_documents",
        "DELETE FROM claims",
        "DELETE FROM packing_list_audit",
        "DELETE FROM packing_list_batches",
        "DELETE FROM packing_lists",
        "DELETE FROM passenger_audit_logs",
        "DELETE FROM preboarding_forms",
        "DELETE FROM passenger_documents",
        "DELETE FROM passenger_payments",
        "DELETE FROM passengers",
        "DELETE FROM passenger_bookings",
        "DELETE FROM rate_offers",
        "DELETE FROM rate_grid_lines",
        "DELETE FROM rate_grids",
        "DELETE FROM order_assignments",
        "DELETE FROM orders",
        "DELETE FROM clients",
        "DELETE FROM crew_tickets",
        "DELETE FROM crew_assignments",
        "DELETE FROM sof_events",
        "DELETE FROM cargo_document_attachments",
        "DELETE FROM cargo_documents",
        "DELETE FROM onboard_attachments",
        "DELETE FROM onboard_notifications",
        "DELETE FROM eta_shifts",
        "DELETE FROM docker_shifts",
        "DELETE FROM escale_operations",
        "DELETE FROM leg_finances",
        "DELETE FROM leg_kpis",
        "DELETE FROM planning_shares",
        "DELETE FROM vessel_positions",
        "DELETE FROM legs",
    ]
    for sql in purge_sql:
        try:
            await db.execute(text(sql))
        except Exception as e:
            logger.warning(f"Full purge SQL failed ({sql[:50]}): {e}")
    await log_activity(db, user=user, action="reset", module="admin",
                       detail="Full database reset",
                       ip_address=get_client_ip(request))
    await db.flush()
    return RedirectResponse(url="/admin/settings#database", status_code=303)


@router.get("/database/stats", response_class=HTMLResponse)
async def database_stats(
    request: Request,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    from sqlalchemy import text
    result = await db.execute(text(
        "SELECT tablename FROM pg_tables WHERE schemaname = 'public' ORDER BY tablename"
    ))
    tables = [r[0] for r in result.fetchall()]
    stats = []
    for t in tables:
        if t not in ALLOWED_TABLES:
            continue
        try:
            cnt = await db.execute(text(f'SELECT COUNT(*) FROM "{t}"'))
            count = cnt.scalar() or 0
            stats.append((t, count))
        except Exception as e:
            logger.warning(f"DB stats query failed for {t}: {e}")
            stats.append((t, "?"))

    html = '<div class="db-stat-grid">'
    for name, count in stats:
        html += f'<div class="db-stat-item"><span>{name}</span><strong>{count}</strong></div>'
    html += '</div>'
    return HTMLResponse(html)


@router.post("/database/cleanup-notifications", response_class=HTMLResponse)
async def cleanup_notifications(
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    from sqlalchemy import text
    await db.execute(text(
        "DELETE FROM notifications WHERE is_archived = TRUE AND created_at < NOW() - INTERVAL '30 days'"
    ))
    await db.flush()
    return RedirectResponse(url="/admin/settings#database", status_code=303)


@router.post("/database/cleanup-audit", response_class=HTMLResponse)
async def cleanup_audit(
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    from sqlalchemy import text
    await db.execute(text(
        "DELETE FROM packing_list_audit WHERE created_at < NOW() - INTERVAL '12 months'"
    ))
    await db.execute(text(
        "DELETE FROM passenger_audit_logs WHERE created_at < NOW() - INTERVAL '12 months'"
    ))
    await db.flush()
    return RedirectResponse(url="/admin/settings#database", status_code=303)


@router.post("/database/cleanup-activity-logs", response_class=HTMLResponse)
async def cleanup_activity_logs(
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    from sqlalchemy import text
    await db.execute(text(
        "DELETE FROM activity_logs WHERE created_at < NOW() - INTERVAL '6 months'"
    ))
    await db.flush()
    return RedirectResponse(url="/admin/settings#database", status_code=303)


@router.post("/database/cleanup-onboard-notifications", response_class=HTMLResponse)
async def cleanup_onboard_notifications(
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    from sqlalchemy import text
    await db.execute(text(
        "DELETE FROM onboard_notifications WHERE is_read = TRUE AND created_at < NOW() - INTERVAL '30 days'"
    ))
    await db.flush()
    return RedirectResponse(url="/admin/settings#database", status_code=303)


@router.post("/database/cleanup-access-logs", response_class=HTMLResponse)
async def cleanup_access_logs(
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    from sqlalchemy import text
    await db.execute(text(
        "DELETE FROM portal_access_logs WHERE accessed_at < NOW() - INTERVAL '6 months'"
    ))
    await db.flush()
    return RedirectResponse(url="/admin/settings#database", status_code=303)


# ─── USER LANGUAGE ────────────────────────────────────────────
@router.post("/settings/language", response_class=HTMLResponse)
async def update_language(
    request: Request,
    language: str = Form(...),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    from app.i18n import SUPPORTED_LANGUAGES
    if language in SUPPORTED_LANGUAGES:
        user.language = language
        await db.flush()
    url = "/admin/settings"
    response = RedirectResponse(url=url, status_code=303)
    response.set_cookie("towt_lang", language, max_age=365*24*3600,
                        httponly=True, samesite="lax", secure=True)
    return response


# ─── CABIN PRICING GRID ──────────────────────────────────────
@account_router.post("/settings/pricing/add", response_class=HTMLResponse)
async def pricing_add(
    request: Request,
    origin_locode: str = Form(...),
    destination_locode: str = Form(...),
    cabin_type: str = Form(...),
    price: float = Form(...),
    deposit_pct: int = Form(30),
    notes: str = Form(""),
    user: User = Depends(require_admin_or_commercial),
    db: AsyncSession = Depends(get_db),
):
    from app.models.passenger import CabinPriceGrid
    entry = CabinPriceGrid(
        origin_locode=origin_locode.strip().upper(),
        destination_locode=destination_locode.strip().upper(),
        cabin_type=cabin_type,
        price=price,
        deposit_pct=deposit_pct,
        notes=notes.strip() or None,
    )
    db.add(entry)
    await db.flush()
    return RedirectResponse(url="/admin/settings#pricing", status_code=303)


@account_router.post("/settings/pricing/{price_id}/edit", response_class=HTMLResponse)
async def pricing_edit(
    price_id: int, request: Request,
    price: float = Form(...),
    deposit_pct: int = Form(30),
    notes: str = Form(""),
    is_active: Optional[str] = Form(None),
    user: User = Depends(require_admin_or_commercial),
    db: AsyncSession = Depends(get_db),
):
    from app.models.passenger import CabinPriceGrid
    entry = await db.get(CabinPriceGrid, price_id)
    if entry:
        entry.price = price
        entry.deposit_pct = deposit_pct
        entry.notes = notes.strip() or None
        entry.is_active = is_active == "on"
        await db.flush()
    return RedirectResponse(url="/admin/settings#pricing", status_code=303)


@account_router.delete("/settings/pricing/{price_id}", response_class=HTMLResponse)
async def pricing_delete(
    price_id: int, request: Request,
    user: User = Depends(require_admin_or_commercial),
    db: AsyncSession = Depends(get_db),
):
    from app.models.passenger import CabinPriceGrid
    entry = await db.get(CabinPriceGrid, price_id)
    if entry:
        await db.delete(entry)
        await db.flush()
    return HTMLResponse("", status_code=200)


# ─── INSURANCE CONTRACTS ─────────────────────────────────────
@router.post("/settings/insurance/add", response_class=HTMLResponse)
async def insurance_add(
    request: Request,
    guarantee_type: str = Form(...),
    insurer_name: str = Form(...),
    insurer_email: str = Form(""),
    insurer_phone: str = Form(""),
    insurer_address: str = Form(""),
    broker_reference: str = Form(""),
    franchise_amount: str = Form("0"),
    guarantee_ceiling: str = Form("0"),
    policy_number: str = Form(""),
    notes: str = Form(""),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    from app.models.finance import InsuranceContract
    entry = InsuranceContract(
        guarantee_type=guarantee_type,
        insurer_name=insurer_name,
        insurer_email=insurer_email or None,
        insurer_phone=insurer_phone or None,
        insurer_address=insurer_address or None,
        broker_reference=broker_reference or None,
        franchise_amount=pf(franchise_amount, 0),
        guarantee_ceiling=pf(guarantee_ceiling, 0),
        policy_number=policy_number or None,
        notes=notes or None,
    )
    db.add(entry)
    await db.flush()
    return RedirectResponse(url="/admin/settings#insurance", status_code=303)


@router.post("/settings/insurance/{ins_id}/edit", response_class=HTMLResponse)
async def insurance_edit(
    ins_id: int, request: Request,
    insurer_name: str = Form(...),
    insurer_email: str = Form(""),
    insurer_phone: str = Form(""),
    insurer_address: str = Form(""),
    broker_reference: str = Form(""),
    franchise_amount: str = Form("0"),
    guarantee_ceiling: str = Form("0"),
    policy_number: str = Form(""),
    is_active: Optional[str] = Form(None),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    from app.models.finance import InsuranceContract
    entry = await db.get(InsuranceContract, ins_id)
    if entry:
        entry.insurer_name = insurer_name
        entry.insurer_email = insurer_email or None
        entry.insurer_phone = insurer_phone or None
        entry.insurer_address = insurer_address or None
        entry.broker_reference = broker_reference or None
        entry.franchise_amount = pf(franchise_amount, 0)
        entry.guarantee_ceiling = pf(guarantee_ceiling, 0)
        entry.policy_number = policy_number or None
        entry.is_active = is_active == "on"
        await db.flush()
    return RedirectResponse(url="/admin/settings#insurance", status_code=303)


# ─── PIPEDRIVE CRM SETTINGS ──────────────────────────────────
@router.post("/settings/pipedrive/update", response_class=HTMLResponse)
async def pipedrive_update(
    request: Request,
    pipedrive_api_token: str = Form(""),
    user: User = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    """Save Pipedrive API token in database."""
    token_val = pipedrive_api_token.strip()
    result = await db.execute(
        select(OpexParameter).where(OpexParameter.parameter_name == "pipedrive_api_token")
    )
    param = result.scalar_one_or_none()
    if param:
        param.description = token_val
        param.parameter_value = 1 if token_val else 0
    else:
        param = OpexParameter(
            parameter_name="pipedrive_api_token",
            parameter_value=1 if token_val else 0,
            unit="",
            category="integrations",
            description=token_val,
        )
        db.add(param)
    await db.flush()
    await log_activity(db, user=user, action="update", module="admin",
                       entity_type="settings", entity_id=None,
                       entity_label="Pipedrive API token updated", ip_address=get_client_ip(request))
    return RedirectResponse(url="/admin/settings#pipedrive", status_code=303)


@router.get("/settings/pipedrive/test")
async def pipedrive_test(
    request: Request,
    user: User = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    """Test Pipedrive API connection."""
    from fastapi.responses import JSONResponse
    from app.utils.pipedrive import _get_token_from_db, _request

    token = await _get_token_from_db(db)
    if not token:
        return JSONResponse(content={"success": False, "error": "Token API non configuré"})

    try:
        import httpx
        async with httpx.AsyncClient(timeout=15.0) as client:
            resp = await client.get(
                "https://api.pipedrive.com/v1/users/me",
                params={"api_token": token},
            )
        if resp.status_code == 401:
            return JSONResponse(content={"success": False, "error": "Token invalide (401 Unauthorized)"})
        if resp.status_code != 200:
            return JSONResponse(content={"success": False, "error": f"Erreur HTTP {resp.status_code}"})
        data = resp.json()
        if not data.get("success"):
            return JSONResponse(content={"success": False, "error": "Token invalide — Pipedrive a rejeté la requête"})
        user_data = data.get("data", {})
        company = user_data.get("company_name", "") or ""
        name = user_data.get("name", "") or ""
        return JSONResponse(content={"success": True, "company": company, "user": name})
    except httpx.ConnectError:
        return JSONResponse(content={"success": False, "error": "Impossible de joindre api.pipedrive.com — vérifiez l'accès réseau du serveur (DNS, firewall, proxy)"})
    except httpx.TimeoutException:
        return JSONResponse(content={"success": False, "error": "Timeout — api.pipedrive.com ne répond pas (>15s)"})
    except Exception as e:
        return JSONResponse(content={"success": False, "error": f"Erreur : {type(e).__name__} — {e}"})


# ─── MON COMPTE (accessible à tous les rôles) ─────────────────
@account_router.get("/my-account", response_class=HTMLResponse)
async def my_account(
    request: Request,
    success: Optional[str] = Query(None),
    error: Optional[str] = Query(None),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    from app.i18n import SUPPORTED_LANGUAGES
    return templates.TemplateResponse("admin/my_account.html", {
        "request": request,
        "user": user,
        "languages": SUPPORTED_LANGUAGES,
        "active_module": "my-account",
        "success": success,
        "error": error,
    })


@account_router.post("/my-account/password", response_class=HTMLResponse)
async def my_account_password(
    request: Request,
    current_password: str = Form(...),
    new_password: str = Form(...),
    confirm_password: str = Form(...),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    from app.auth import verify_password, hash_password
    if not verify_password(current_password, user.hashed_password):
        return RedirectResponse(url="/admin/my-account?error=Mot+de+passe+actuel+incorrect", status_code=303)
    if new_password != confirm_password:
        return RedirectResponse(url="/admin/my-account?error=Les+mots+de+passe+ne+correspondent+pas", status_code=303)
    if len(new_password) < 4:
        return RedirectResponse(url="/admin/my-account?error=Le+mot+de+passe+doit+faire+au+moins+4+caractères", status_code=303)
    user.hashed_password = hash_password(new_password)
    await db.flush()
    return RedirectResponse(url="/admin/my-account?success=Mot+de+passe+modifié+avec+succès", status_code=303)


@account_router.post("/my-account/language", response_class=HTMLResponse)
async def my_account_language(
    request: Request,
    language: str = Form(...),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    from app.i18n import SUPPORTED_LANGUAGES
    if language in SUPPORTED_LANGUAGES:
        user.language = language
        await db.flush()
    return RedirectResponse(url="/admin/my-account?success=Langue+modifiée+avec+succès", status_code=303)


# ─── ACTIVITY LOGS ──────────────────────────────────────────
@router.get("/activity-logs", response_class=HTMLResponse)
async def activity_logs(
    request: Request,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
    page: int = Query(1, ge=1),
    action: Optional[str] = Query(None),
    module: Optional[str] = Query(None),
    user_id: Optional[int] = Query(None),
):
    from app.models.activity_log import ActivityLog
    from sqlalchemy import desc

    per_page = 50
    q = select(ActivityLog)

    if action:
        q = q.where(ActivityLog.action == action)
    if module:
        q = q.where(ActivityLog.module == module)
    if user_id:
        q = q.where(ActivityLog.user_id == user_id)

    # Count
    count_q = select(func.count()).select_from(q.subquery())
    total = (await db.execute(count_q)).scalar() or 0
    total_pages = max(1, (total + per_page - 1) // per_page)

    # Fetch page
    q = q.order_by(desc(ActivityLog.created_at)).offset((page - 1) * per_page).limit(per_page)
    result = await db.execute(q)
    logs = result.scalars().all()

    # Action labels
    action_labels = {
        "login": "🔑 Connexion",
        "logout": "🚪 Déconnexion",
        "login_fail": "🚫 Échec",
        "create": "➕ Création",
        "update": "✏️ Modification",
        "delete": "🗑️ Suppression",
    }

    # Build HTML
    rows = []
    for log in logs:
        ts = log.created_at.strftime("%d/%m/%Y %H:%M") if log.created_at else ""
        a_label = action_labels.get(log.action, log.action)
        a_cls = f"al-{log.action}"

        entity = ""
        if log.entity_type:
            entity = f'<span style="font-size:11px;color:#888;">{log.entity_type}</span>'
            if log.entity_label:
                entity += f' <strong style="font-size:12px;">{log.entity_label}</strong>'

        detail_html = ""
        if log.detail:
            detail_html = f'<div style="font-size:10px;color:#888;margin-top:2px;">{log.detail}</div>'

        rows.append(f'''<tr>
            <td class="al-time">{ts}</td>
            <td>{log.user_name or '<span style="color:#ccc;">—</span>'}</td>
            <td><span class="al-action {a_cls}">{a_label}</span></td>
            <td><span class="al-module">{log.module}</span></td>
            <td>{entity}{detail_html}</td>
            <td class="al-ip">{log.ip_address or ''}</td>
        </tr>''')

    # Pagination
    pag = ""
    if total_pages > 1:
        pag = '<div style="display:flex;gap:4px;justify-content:center;margin-top:12px;">'
        if page > 1:
            pag += f'<button class="btn btn-sm btn-secondary" onclick="loadActivityLogs({page-1})"><i data-lucide="chevron-left"></i></button>'
        pag += f'<span style="padding:6px 12px;font-size:12px;color:#666;">Page {page}/{total_pages} ({total} entrées)</span>'
        if page < total_pages:
            pag += f'<button class="btn btn-sm btn-secondary" onclick="loadActivityLogs({page+1})"><i data-lucide="chevron-right"></i></button>'
        pag += '</div>'

    if not rows:
        html = '<div style="text-align:center;padding:40px;color:#888;"><i data-lucide="clipboard-list" style="width:32px;height:32px;"></i><p style="margin-top:8px;">Aucune activité enregistrée</p></div>'
    else:
        html = f'''<div class="table-container"><table class="data-table">
            <thead><tr>
                <th>Date</th><th>Utilisateur</th><th>Action</th><th>Module</th><th>Détails</th><th>IP</th>
            </tr></thead>
            <tbody>{"".join(rows)}</tbody>
        </table></div>{pag}'''

    return HTMLResponse(content=html)


# ─── CO2 DECARBONATION VARIABLES ────────────────────────────
@router.post("/co2/update", response_class=HTMLResponse)
async def update_co2_variables(
    request: Request,
    user: User = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    form = await request.form()
    from datetime import date
    today = date.today()
    for var_name, info in CO2_DEFAULTS.items():
        form_val = form.get(f"co2_{var_name}", "")
        if form_val:
            new_val = pf(form_val, info["value"])
            result = await db.execute(
                select(Co2Variable).where(Co2Variable.variable_name == var_name, Co2Variable.is_current == True)
            )
            existing = result.scalar_one_or_none()
            if existing and abs(existing.variable_value - new_val) > 0.0001:
                # Historize TOWT CO2 EF
                if var_name == "towt_co2_ef":
                    existing.is_current = False
                    new_var = Co2Variable(
                        variable_name=var_name, variable_value=new_val,
                        unit=info["unit"], description=info.get("description", ""),
                        effective_date=today, is_current=True,
                    )
                    db.add(new_var)
                else:
                    existing.variable_value = new_val
            elif not existing:
                db.add(Co2Variable(
                    variable_name=var_name, variable_value=new_val,
                    unit=info["unit"], description=info.get("description", ""),
                    effective_date=today, is_current=True,
                ))
    await db.flush()
    await log_activity(db, user, "admin", "update", "Co2Variable", None, "Mise à jour variables CO2")
    url = "/admin/settings"
    if request.headers.get("HX-Request"):
        return HTMLResponse(content="", headers={"HX-Redirect": url})
    return RedirectResponse(url=url, status_code=303)


# ─── MRV PARAMETERS ─────────────────────────────────────────
@router.post("/mrv/update", response_class=HTMLResponse)
async def update_mrv_parameters(
    request: Request,
    user: User = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    form = await request.form()
    for param_name, info in MRV_DEFAULTS.items():
        form_val = form.get(f"mrv_{param_name}", "")
        if form_val:
            new_val = pf(form_val, info["value"])
            result = await db.execute(
                select(MrvParameter).where(MrvParameter.parameter_name == param_name)
            )
            existing = result.scalar_one_or_none()
            if existing:
                existing.parameter_value = new_val
            else:
                db.add(MrvParameter(
                    parameter_name=param_name, parameter_value=new_val,
                    unit=info["unit"], description=info.get("description", ""),
                ))
    await db.flush()
    await log_activity(db, user, "admin", "update", "MrvParameter", None, "Mise à jour paramètres MRV")
    url = "/admin/settings"
    if request.headers.get("HX-Request"):
        return HTMLResponse(content="", headers={"HX-Redirect": url})
    return RedirectResponse(url=url, status_code=303)


# ─── EMISSIONS UPDATE ────────────────────────────────────────
@router.post("/emissions/update", response_class=HTMLResponse)
async def update_emission_params(
    request: Request,
    user: User = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    form = await request.form()
    for key in form.keys():
        if key.startswith("ep_"):
            param_name = key[3:]
            new_val = pf(form.get(key), 0)
            result = await db.execute(
                select(EmissionParameter).where(EmissionParameter.parameter_name == param_name)
            )
            ep = result.scalar_one_or_none()
            if ep:
                ep.parameter_value = new_val
    await db.flush()
    await log_activity(db, user, "admin", "update", "EmissionParameter", None, "Mise à jour émissions KPI")
    url = "/admin/settings"
    if request.headers.get("HX-Request"):
        return HTMLResponse(content="", headers={"HX-Redirect": url})
    return RedirectResponse(url=url, status_code=303)


# ─── IMPORT PLANNING CSV ────────────────────────────────────
# Port coordinates for auto-creation
_PORT_COORDS = {
    "FRFEC": ("Fecamp", 49.7578, 0.3650, "FR"),
    "FRLEH": ("Le Havre", 49.4944, 0.1079, "FR"),
    "FRCOC": ("Concarneau", 47.8706, -3.9214, "FR"),
    "USNYC": ("New York", 40.6892, -74.0445, "US"),
    "BRSSO": ("Sao Sebastiao", -23.8159, -45.4097, "BR"),
    "COSMR": ("Santa Marta", 11.2472, -74.1992, "CO"),
    "COSTM": ("Santa Marta", 11.2472, -74.1992, "CO"),
    "HNPCR": ("Puerto Cortes", 15.8383, -87.9550, "HN"),
    "CAQUE": ("Quebec", 46.8139, -71.2080, "CA"),
    "CAMAT": ("Matane", 48.8500, -67.5333, "CA"),
    "GPPTP": ("Pointe-a-Pitre", 16.2411, -61.5310, "GP"),
    "GTSTC": ("Puerto Santo Tomas de Castilla", 15.7000, -88.6167, "GT"),
    "PTOPO": ("Porto", 41.1496, -8.6109, "PT"),
    "CUHAV": ("La Habana", 23.1136, -82.3666, "CU"),
    "VNSGN": ("Ho Chi Minh City", 10.7626, 106.7533, "VN"),
    "VNDAD": ("Da Nang", 16.0678, 108.2208, "VN"),
    "REPDG": ("Port de Pointe des Galets", -20.9342, 55.2917, "RE"),
    "RELPT": ("Le Port", -20.9342, 55.2917, "RE"),
}


def _parse_dt(s: str):
    """Parse DD/MM/YYYY HH:MM or DD/MM/YYYY HH:MM:SS."""
    if not s or not s.strip():
        return None
    s = s.strip()
    for fmt in ("%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


@router.post("/import/planning", response_class=HTMLResponse)
async def import_planning_csv(
    request: Request,
    file: UploadFile = File(...),
    user: User = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    """Import a planning CSV file (semicolon-separated) with legs data."""
    from app.models.port import Port

    raw = await file.read()
    try:
        content = raw.decode("utf-8")
    except UnicodeDecodeError:
        content = raw.decode("latin-1")

    reader = csv.DictReader(io.StringIO(content), delimiter=";")
    rows = list(reader)

    if not rows:
        return RedirectResponse(url="/admin/settings?import_error=Fichier+vide", status_code=303)

    # Build vessel lookup {name -> id}
    v_result = await db.execute(select(Vessel))
    vessel_map = {v.name.lower(): v for v in v_result.scalars().all()}

    created = 0
    updated = 0
    ports_created = 0
    errors = []

    for i, row in enumerate(rows, 1):
        try:
            leg_code = row.get("Leg Code", "").strip()
            vessel_name = row.get("Navire", "").strip()
            dep_locode = row.get("Départ LOCODE", row.get("Depart LOCODE", "")).strip()
            arr_locode = row.get("Arrivée LOCODE", row.get("Arrivee LOCODE", "")).strip()
            dep_name = row.get("Port Départ", row.get("Port Depart", "")).strip()
            arr_name = row.get("Port Arrivée", row.get("Port Arrivee", "")).strip()

            if not leg_code or not vessel_name:
                continue

            # Find vessel
            vessel = vessel_map.get(vessel_name.lower())
            if not vessel:
                errors.append(f"Ligne {i}: navire '{vessel_name}' introuvable")
                continue

            # Ensure departure port exists
            for locode, name in [(dep_locode, dep_name), (arr_locode, arr_name)]:
                p_result = await db.execute(select(Port).where(Port.locode == locode))
                port = p_result.scalar_one_or_none()
                if not port:
                    coords = _PORT_COORDS.get(locode)
                    port = Port(
                        locode=locode,
                        name=coords[0] if coords else name,
                        latitude=coords[1] if coords else None,
                        longitude=coords[2] if coords else None,
                        country_code=coords[3] if coords else locode[:2],
                    )
                    db.add(port)
                    await db.flush()
                    ports_created += 1
                elif not port.latitude and locode in _PORT_COORDS:
                    port.latitude = _PORT_COORDS[locode][1]
                    port.longitude = _PORT_COORDS[locode][2]

            # Parse dates
            etd = _parse_dt(row.get("ETD", ""))
            eta = _parse_dt(row.get("ETA", ""))
            atd = _parse_dt(row.get("ATD (réel)", row.get("ATD (reel)", "")))
            ata = _parse_dt(row.get("ATA (réel)", row.get("ATA (reel)", "")))

            # Parse numbers
            distance_ortho = pf(row.get("Distance Ortho (NM)", ""), None)
            distance_reel = pf(row.get("Distance Réelle (NM)", row.get("Distance Reelle (NM)", "")), None)
            speed = pf(row.get("Vitesse (nds)", ""), None)
            duration = pf(row.get("Durée Est. (h)", row.get("Duree Est. (h)", "")), None)
            status = row.get("Statut", "planned").strip() or "planned"

            # Extract year from last digit of leg_code
            last_digit = leg_code.rstrip("-ABCDEFGHIJKLMNOPQRSTUVWXYZ")[-1:]
            year_map = {"4": 2024, "5": 2025, "6": 2026, "7": 2027}
            year = year_map.get(last_digit, etd.year if etd else 2026)

            # Extract sequence from second character (letter)
            seq_char = leg_code[1] if len(leg_code) > 1 else "A"
            sequence = ord(seq_char.upper()) - ord('A') + 1 if seq_char.isalpha() else 1

            # Compute elongation
            elongation = None
            if distance_ortho and distance_reel and distance_ortho > 0:
                elongation = round(distance_reel / distance_ortho, 2)

            # Upsert leg
            result = await db.execute(select(Leg).where(Leg.leg_code == leg_code))
            leg = result.scalar_one_or_none()

            if leg:
                leg.vessel_id = vessel.id
                leg.year = year
                leg.sequence = sequence
                leg.departure_port_locode = dep_locode
                leg.arrival_port_locode = arr_locode
                leg.etd = etd
                leg.eta = eta
                leg.atd = atd
                leg.ata = ata
                leg.distance_nm = distance_ortho
                leg.computed_distance = distance_reel
                leg.speed_knots = speed
                leg.estimated_duration_hours = duration
                leg.elongation_coeff = elongation or leg.elongation_coeff
                leg.status = status
                updated += 1
            else:
                leg = Leg(
                    leg_code=leg_code,
                    vessel_id=vessel.id,
                    year=year,
                    sequence=sequence,
                    departure_port_locode=dep_locode,
                    arrival_port_locode=arr_locode,
                    etd=etd, eta=eta, atd=atd, ata=ata,
                    distance_nm=distance_ortho,
                    computed_distance=distance_reel,
                    speed_knots=speed,
                    elongation_coeff=elongation or 1.25,
                    estimated_duration_hours=duration,
                    status=status,
                )
                db.add(leg)
                created += 1

        except Exception as e:
            errors.append(f"Ligne {i}: {str(e)}")

    await db.flush()
    await log_activity(db, user=user, action="import", module="admin",
                       entity_type="planning_csv", entity_id=None,
                       entity_label=f"{file.filename}: {created} créés, {updated} mis à jour",
                       ip_address=get_client_ip(request))

    msg = f"Import terminé : {created} legs créés, {updated} mis à jour, {ports_created} ports créés"
    if errors:
        msg += f", {len(errors)} erreurs"
    url = f"/admin/settings?import_success={msg.replace(' ', '+')}"
    if request.headers.get("HX-Request"):
        return HTMLResponse(content="", headers={"HX-Redirect": url})
    return RedirectResponse(url=url, status_code=303)


# ═══ RGPD DATA EXPORT (DSR) ═════════════════════════════════
@router.get("/rgpd/export/passenger/{booking_id}")
async def rgpd_export_passenger(
    booking_id: int, request: Request,
    user: User = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    """Export all personal data for a passenger booking (GDPR Data Subject Request)."""
    import json
    from app.models.passenger import PassengerBooking, Passenger, PreBoardingForm
    from app.models.portal_message import PortalMessage
    from app.models.portal_access_log import PortalAccessLog

    result = await db.execute(
        select(PassengerBooking).options(
            selectinload(PassengerBooking.passengers),
            selectinload(PassengerBooking.payments),
        ).where(PassengerBooking.id == booking_id)
    )
    booking = result.scalar_one_or_none()
    if not booking:
        raise HTTPException(404)

    data = {
        "export_type": "RGPD_DSR",
        "export_date": datetime.now().isoformat(),
        "exported_by": user.username,
        "booking": {
            "reference": booking.reference,
            "status": booking.status,
            "booking_date": str(booking.booking_date) if booking.booking_date else None,
            "contact_email": booking.contact_email,
            "contact_phone": booking.contact_phone,
            "created_at": booking.created_at.isoformat() if booking.created_at else None,
        },
        "passengers": [],
        "payments": [],
        "messages": [],
        "access_logs": [],
    }

    for pax in booking.passengers:
        pax_data = {
            "first_name": pax.first_name, "last_name": pax.last_name,
            "email": pax.email, "phone": pax.phone,
            "date_of_birth": str(pax.date_of_birth) if pax.date_of_birth else None,
            "nationality": pax.nationality, "passport_number": pax.passport_number,
            "emergency_contact_name": pax.emergency_contact_name,
            "emergency_contact_phone": pax.emergency_contact_phone,
        }
        # Pre-boarding form
        form_result = await db.execute(select(PreBoardingForm).where(PreBoardingForm.passenger_id == pax.id))
        form = form_result.scalar_one_or_none()
        if form:
            pax_data["questionnaire"] = {
                "sailed_before": form.sailed_before, "seasick": form.seasick,
                "chronic_conditions": form.chronic_conditions, "allergies": form.allergies,
                "daily_medication": form.daily_medication, "dietary_requirements": form.dietary_requirements,
                "gdpr_consent": form.gdpr_consent,
                "gdpr_consent_at": form.gdpr_consent_at.isoformat() if form.gdpr_consent_at else None,
                "signed_at": form.signed_at.isoformat() if form.signed_at else None,
            }
        data["passengers"].append(pax_data)

    for pay in booking.payments:
        data["payments"].append({
            "type": pay.payment_type, "amount": str(pay.amount),
            "status": pay.status, "paid_date": str(pay.paid_date) if pay.paid_date else None,
        })

    # Messages
    msg_result = await db.execute(
        select(PortalMessage).where(PortalMessage.booking_id == booking_id).order_by(PortalMessage.created_at))
    for msg in msg_result.scalars().all():
        data["messages"].append({
            "sender": msg.sender_name, "message": msg.message,
            "created_at": msg.created_at.isoformat() if msg.created_at else None,
        })

    # Access logs
    log_result = await db.execute(
        select(PortalAccessLog).where(PortalAccessLog.booking_id == booking_id)
        .order_by(PortalAccessLog.accessed_at.desc()).limit(100))
    for log in log_result.scalars().all():
        data["access_logs"].append({
            "ip": log.ip_address, "path": log.path,
            "accessed_at": log.accessed_at.isoformat() if log.accessed_at else None,
        })

    await log_activity(db, user=user, action="rgpd_export", module="admin",
                       entity_type="passenger_booking", entity_id=booking_id,
                       entity_label=f"Export RGPD réservation {booking.reference}",
                       ip_address=get_client_ip(request))

    content = json.dumps(data, indent=2, ensure_ascii=False)
    return StreamingResponse(
        io.BytesIO(content.encode("utf-8")),
        media_type="application/json",
        headers={"Content-Disposition": f"attachment; filename=RGPD_export_{booking.reference}.json"},
    )


# ─── RGPD: RIGHT TO ERASURE (droit à l'effacement) ─────────
@router.post("/rgpd/erase/{booking_id}", response_class=HTMLResponse)
async def rgpd_erase_booking(
    booking_id: int,
    request: Request,
    user: User = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    """Anonymize personal data for a passenger booking (RGPD Art. 17 right to erasure).

    Replaces all personal data with anonymized placeholders while preserving
    booking structure for statistical/financial purposes.
    """
    from app.models.passenger import (
        PassengerBooking, Passenger, PassengerDocument, PassengerPayment,
        PreBoardingForm, PassengerAuditLog,
    )
    import os

    result = await db.execute(
        select(PassengerBooking).options(
            selectinload(PassengerBooking.passengers).selectinload(Passenger.documents),
            selectinload(PassengerBooking.payments),
        ).where(PassengerBooking.id == booking_id)
    )
    booking = result.scalar_one_or_none()
    if not booking:
        raise HTTPException(404, detail="Réservation introuvable")

    anon_label = f"ANONYMIZED-{booking_id}"

    # Anonymize booking contact info
    booking.contact_email = f"{anon_label}@erased.local"
    booking.contact_phone = "ERASED"
    booking.notes = "Données personnelles effacées (RGPD Art. 17)"
    booking.token = f"erased-{booking_id}-{secrets.token_urlsafe(8)}"

    # Anonymize each passenger
    for pax in booking.passengers:
        pax.first_name = "ERASED"
        pax.last_name = anon_label
        pax.email = f"{anon_label}@erased.local"
        pax.phone = "ERASED"
        pax.date_of_birth = None
        pax.nationality = "XX"
        pax.passport_number = "ERASED"
        pax.emergency_contact_name = "ERASED"
        pax.emergency_contact_phone = "ERASED"

        # Delete uploaded documents (files on disk)
        for doc in pax.documents:
            if doc.file_path and os.path.isfile(doc.file_path):
                try:
                    os.remove(doc.file_path)
                except OSError:
                    pass
            doc.file_path = None
            doc.filename = "ERASED"
            doc.status = "missing"

        # Anonymize pre-boarding forms
        form_result = await db.execute(
            select(PreBoardingForm).where(PreBoardingForm.passenger_id == pax.id)
        )
        for form in form_result.scalars().all():
            form.chronic_conditions = "ERASED"
            form.allergies = "ERASED"
            form.daily_medication = "ERASED"
            form.dietary_requirements = "ERASED"
            form.intolerances = "ERASED"

    # Anonymize payment references
    for payment in booking.payments:
        payment.reference = f"ERASED-{payment.id}"
        payment.notes = "ERASED"

    # Delete audit logs for this booking
    audit_result = await db.execute(
        select(PassengerAuditLog).where(PassengerAuditLog.booking_id == booking_id)
    )
    for audit in audit_result.scalars().all():
        audit.user_email = "ERASED"
        audit.detail = "ERASED (RGPD)"

    # Delete portal access logs
    try:
        from app.models.portal_access_log import PortalAccessLog
        access_result = await db.execute(
            select(PortalAccessLog).where(
                PortalAccessLog.token == booking.token,
            )
        )
        for log in access_result.scalars().all():
            log.ip_address = "0.0.0.0"
            log.user_agent = "ERASED"
    except Exception:
        pass  # Table may not exist

    await db.flush()
    await log_activity(db, user=user, action="rgpd_erase", module="admin",
                       entity_type="passenger_booking", entity_id=booking_id,
                       entity_label=f"Effacement RGPD réservation {booking.reference}",
                       ip_address=get_client_ip(request))

    url = "/admin/settings?section=database"
    if request.headers.get("HX-Request"):
        return HTMLResponse(content="", headers={"HX-Redirect": url})
    return RedirectResponse(url=url, status_code=303)
