from fastapi import APIRouter, Request, Depends, Form, HTTPException, Query, UploadFile, File
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse, FileResponse
from app.templating import templates
from sqlalchemy.ext.asyncio import AsyncSession
from app.utils.activity import log_activity, get_client_ip
from sqlalchemy import select, func
from sqlalchemy.orm import selectinload
from typing import Optional
from datetime import datetime, timezone
import io

from app.database import get_db
from app.auth import get_current_user
from app.permissions import require_permission
from app.models.user import User
from app.models.order import Order
from app.models.leg import Leg
from app.models.vessel import Vessel
from app.models.packing_list import PackingList, PackingListBatch, PackingListAudit, PackingListDocument
from app.models.portal_message import PortalMessage
from app.utils.notifications import notify_cargo_progress
from app.routers.kpi_router import compute_decarbonation, get_co2_variables
from app.models.crew import CrewAssignment, CrewMember
from app.i18n import get_lang_from_request

router = APIRouter(prefix="/cargo", tags=["cargo"])

IMO_CLASSES = [
    "Non-Dangerous Goods",
    "Class 1: Explosives",
    "Class 2: Gases",
    "Class 3: Flammable liquids",
    "Class 4: Flammable solids",
    "Class 5: Oxidizing substances",
    "Class 6: Toxic substances",
    "Class 7: Radioactive material",
    "Class 8: Corrosive substances",
    "Class 9: Miscellaneous",
]


def pf(val, default=None):
    if val is None or (isinstance(val, str) and val.strip() == ""):
        return default
    try:
        s = val.replace(" ", "").replace("\u00a0", "")
        if "." in s and "," in s:
            s = s.replace(".", "").replace(",", ".")
        elif "," in s:
            s = s.replace(",", ".")
        return float(s)
    except Exception:
        return default


def pi(val, default=None):
    if val is None or (isinstance(val, str) and val.strip() == ""):
        return default
    try:
        return int(val)
    except Exception:
        return default


# === EXPLOITATION VIEW ===
@router.get("", response_class=HTMLResponse)
@router.get("/", response_class=HTMLResponse)
async def cargo_home(
    request: Request,
    status: Optional[str] = Query(None),
    user: User = Depends(require_permission("cargo", "C")),
    db: AsyncSession = Depends(get_db),
):
    query = (
        select(PackingList)
        .options(
            selectinload(PackingList.order).selectinload(Order.leg).selectinload(Leg.vessel),
            selectinload(PackingList.order).selectinload(Order.leg).selectinload(Leg.departure_port),
            selectinload(PackingList.order).selectinload(Order.leg).selectinload(Leg.arrival_port),
            selectinload(PackingList.batches),
        )
        .order_by(PackingList.created_at.desc())
    )
    if status:
        query = query.where(PackingList.status == status)
    result = await db.execute(query)
    packing_lists = result.scalars().all()

    return templates.TemplateResponse("cargo/index.html", {
        "request": request, "user": user,
        "packing_lists": packing_lists,
        "selected_status": status,
        "active_module": "cargo",
    })


# === CREATE PACKING LIST FROM ORDER ===
@router.post("/create", response_class=HTMLResponse)
async def create_packing_list(
    request: Request,
    order_id: str = Form(...),
    user: User = Depends(require_permission("cargo", "M")),
    db: AsyncSession = Depends(get_db),
):
    oid = int(order_id)
    order_result = await db.execute(
        select(Order).options(
            selectinload(Order.leg).selectinload(Leg.vessel),
            selectinload(Order.leg).selectinload(Leg.departure_port),
            selectinload(Order.leg).selectinload(Leg.arrival_port),
        ).where(Order.id == oid)
    )
    order = order_result.scalar_one_or_none()
    if not order:
        raise HTTPException(404)

    # Check if packing list already exists
    existing = await db.execute(select(PackingList).where(PackingList.order_id == oid))
    if existing.scalar_one_or_none():
        url = "/cargo"
        if request.headers.get("HX-Request"):
            return HTMLResponse(content="", headers={"HX-Redirect": url})
        return RedirectResponse(url=url, status_code=303)

    pl = PackingList(order_id=oid)
    db.add(pl)
    await db.flush()
    await log_activity(db, user=user, action="create", module="cargo",
                       entity_type="packing_list", entity_id=pl.id,
                       entity_label=f"PL for {order.reference}",
                       ip_address=get_client_ip(request))

    # Create one default batch pre-filled with TOWT data
    batch = PackingListBatch(
        packing_list_id=pl.id,
        batch_number=1,
        booking_confirmation=order.reference,
        customer_name=order.client_name,
        freight_rate=order.unit_price,
    )
    if order.leg:
        leg = order.leg
        batch.voyage_id = leg.leg_code
        batch.vessel = leg.vessel.name if leg.vessel else None
        batch.loading_date = leg.etd.date() if leg.etd else None
        batch.pol_code = leg.departure_port.locode if leg.departure_port else None
        batch.pod_code = leg.arrival_port.locode if leg.arrival_port else None
        batch.pol_name = leg.departure_port.name if leg.departure_port else None
        batch.pod_name = leg.arrival_port.name if leg.arrival_port else None

    db.add(batch)
    await db.flush()

    url = "/cargo"
    if request.headers.get("HX-Request"):
        return HTMLResponse(content="", headers={"HX-Redirect": url})
    return RedirectResponse(url=url, status_code=303)


# === VIEW PACKING LIST DETAIL (exploitation) ===
@router.get("/{pl_id}", response_class=HTMLResponse)
async def cargo_detail(
    pl_id: int, request: Request,
    user: User = Depends(require_permission("cargo", "C")),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(
        select(PackingList).options(
            selectinload(PackingList.order).selectinload(Order.leg).selectinload(Leg.vessel),
            selectinload(PackingList.order).selectinload(Order.leg).selectinload(Leg.departure_port),
            selectinload(PackingList.order).selectinload(Order.leg).selectinload(Leg.arrival_port),
            selectinload(PackingList.batches),
        ).where(PackingList.id == pl_id)
    )
    pl = result.scalar_one_or_none()
    if not pl:
        raise HTTPException(404)

    # Load portal messages
    msg_result = await db.execute(
        select(PortalMessage)
        .where(PortalMessage.packing_list_id == pl_id)
        .order_by(PortalMessage.created_at)
    )
    portal_messages = msg_result.scalars().all()
    unread_client_msgs = sum(1 for m in portal_messages if m.sender_type == "client" and not m.is_read)

    # IMDG validation warnings
    imdg_warnings = []
    for batch in pl.batches:
        imdg_warnings.extend(validate_imdg(batch))

    return templates.TemplateResponse("cargo/detail.html", {
        "request": request, "user": user,
        "pl": pl, "imo_classes": IMO_CLASSES,
        "active_module": "cargo",
        "portal_messages": portal_messages,
        "unread_client_msgs": unread_client_msgs,
        "imdg_warnings": imdg_warnings,
    })


# === LOCK / UNLOCK ===
@router.post("/{pl_id}/lock", response_class=HTMLResponse)
async def lock_packing_list(
    pl_id: int, request: Request,
    user: User = Depends(require_permission("cargo", "M")),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(select(PackingList).where(PackingList.id == pl_id))
    pl = result.scalar_one_or_none()
    if not pl:
        raise HTTPException(404)
    pl.status = "locked"
    pl.locked_at = datetime.now(timezone.utc)
    pl.locked_by = user.full_name or user.username
    await db.flush()
    url = f"/cargo/{pl_id}"
    if request.headers.get("HX-Request"):
        return HTMLResponse(content="", headers={"HX-Redirect": url})
    return RedirectResponse(url=url, status_code=303)


@router.post("/{pl_id}/unlock", response_class=HTMLResponse)
async def unlock_packing_list(
    pl_id: int, request: Request,
    user: User = Depends(require_permission("cargo", "M")),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(select(PackingList).where(PackingList.id == pl_id))
    pl = result.scalar_one_or_none()
    if not pl:
        raise HTTPException(404)
    pl.status = "draft"
    pl.locked_at = None
    pl.locked_by = None
    await db.flush()
    url = f"/cargo/{pl_id}"
    if request.headers.get("HX-Request"):
        return HTMLResponse(content="", headers={"HX-Redirect": url})
    return RedirectResponse(url=url, status_code=303)


# === DELETE PACKING LIST ===
@router.delete("/{pl_id}", response_class=HTMLResponse)
async def delete_packing_list(
    pl_id: int, request: Request,
    user: User = Depends(require_permission("cargo", "S")),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(select(PackingList).where(PackingList.id == pl_id))
    pl = result.scalar_one_or_none()
    if not pl:
        raise HTTPException(404)
    await db.delete(pl)
    await db.flush()
    url = "/cargo"
    if request.headers.get("HX-Request"):
        return HTMLResponse(content="", headers={"HX-Redirect": url})
    return RedirectResponse(url=url, status_code=303)


# === EXCEL EXPORT ===
@router.get("/{pl_id}/excel")
async def export_excel(
    pl_id: int,
    user: User = Depends(require_permission("cargo", "C")),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(
        select(PackingList).options(
            selectinload(PackingList.order),
            selectinload(PackingList.batches),
        ).where(PackingList.id == pl_id)
    )
    pl = result.scalar_one_or_none()
    if not pl:
        raise HTTPException(404)

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "LOADING_DATA"

    headers = [
        "VOYAGE_ID", "VESSEL", "LOADING_DATE", "POL_CODE", "POD_CODE",
        "POL_NAME", "POD_NAME", "CUSTOMER_NAME", "BOOKING_CONFIRMATION_TOWT",
        "FREIGHT_RATE_PER_PALLET_EURO", "CARGO_VALUE_USD", "FREIGHT_FORWARDER",
        "CODE_TRANSITAIRE", "SHIPPER_NAME", "WH_REFERENCES_SKU", "PO_NUMBER",
        "ADDITIONAL_REFERENCES", "CUSTOMER_BATCH_ID", "BILL_OF_LADING_ID",
        "NOTIFY_ADRESS", "CONSIGNEE_ORDER_ADRESS", "PALLET_TYPE",
        "TYPE_OF_GOODS", "BIO_PRODUCTS", "CASES_QUANTITY", "UNITS_PER_CASE",
        "IMO_PRODUCT_CLASS", "UN_NUMBER", "PALLET_QUANTITY_PER_BATCH",
        "STACKABLE", "HOLD", "LENGTH_CM", "WIDTH_CM", "HEIGHT_CM",
        "WEIGHT_KG", "SURFACE_M2", "VOLUME_M3",
    ]

    header_fill = PatternFill(start_color="095561", end_color="095561", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    yellow_fill = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    yellow_cols = {
        "CUSTOMER_NAME", "FREIGHT_FORWARDER", "CODE_TRANSITAIRE", "SHIPPER_NAME",
        "PO_NUMBER", "CUSTOMER_BATCH_ID", "NOTIFY_ADRESS", "CONSIGNEE_ORDER_ADRESS",
        "PALLET_TYPE", "TYPE_OF_GOODS", "BIO_PRODUCTS", "CASES_QUANTITY",
        "UNITS_PER_CASE", "IMO_PRODUCT_CLASS", "PALLET_QUANTITY_PER_BATCH",
        "LENGTH_CM", "WIDTH_CM", "HEIGHT_CM", "WEIGHT_KG", "CARGO_VALUE_USD",
    }

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = max(15, len(h) + 2)

    for row_idx, batch in enumerate(pl.batches, 2):
        data = {
            "VOYAGE_ID": batch.voyage_id,
            "VESSEL": batch.vessel,
            "LOADING_DATE": batch.loading_date.isoformat() if batch.loading_date else "",
            "POL_CODE": batch.pol_code,
            "POD_CODE": batch.pod_code,
            "POL_NAME": batch.pol_name,
            "POD_NAME": batch.pod_name,
            "CUSTOMER_NAME": batch.customer_name,
            "BOOKING_CONFIRMATION_TOWT": batch.booking_confirmation,
            "FREIGHT_RATE_PER_PALLET_EURO": batch.freight_rate,
            "CARGO_VALUE_USD": batch.cargo_value_usd,
            "FREIGHT_FORWARDER": batch.freight_forwarder,
            "CODE_TRANSITAIRE": batch.code_transitaire,
            "SHIPPER_NAME": batch.shipper_name,
            "WH_REFERENCES_SKU": batch.wh_references_sku,
            "PO_NUMBER": batch.po_number,
            "ADDITIONAL_REFERENCES": batch.additional_references,
            "CUSTOMER_BATCH_ID": batch.customer_batch_id,
            "BILL_OF_LADING_ID": batch.bill_of_lading_id,
            "NOTIFY_ADRESS": batch.notify_address,
            "CONSIGNEE_ORDER_ADRESS": batch.consignee_address,
            "PALLET_TYPE": batch.pallet_type,
            "TYPE_OF_GOODS": batch.type_of_goods,
            "BIO_PRODUCTS": batch.bio_products,
            "CASES_QUANTITY": batch.cases_quantity,
            "UNITS_PER_CASE": batch.units_per_case,
            "IMO_PRODUCT_CLASS": batch.imo_product_class,
            "UN_NUMBER": batch.un_number,
            "PALLET_QUANTITY_PER_BATCH": batch.pallet_quantity,
            "STACKABLE": batch.stackable,
            "HOLD": batch.hold,
            "LENGTH_CM": batch.length_cm,
            "WIDTH_CM": batch.width_cm,
            "HEIGHT_CM": batch.height_cm,
            "WEIGHT_KG": batch.weight_kg,
            "SURFACE_M2": batch.surface_m2,
            "VOLUME_M3": batch.volume_m3,
        }
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=data.get(h, ""))
            cell.border = thin_border
            if h in yellow_cols:
                cell.fill = yellow_fill

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"PACKING_LIST_{pl.order.reference}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# === BILL OF LADING (DOCX) — one per batch ===
@router.get("/{pl_id}/bol/{batch_id}")
async def bill_of_lading_batch(
    pl_id: int, batch_id: int, request: Request,
    user: User = Depends(require_permission("cargo", "C")),
    db: AsyncSession = Depends(get_db),
):
    """Generate a Bill of Lading DOCX for a single batch."""
    result = await db.execute(
        select(PackingList).options(
            selectinload(PackingList.order).selectinload(Order.leg).selectinload(Leg.vessel),
            selectinload(PackingList.order).selectinload(Order.leg).selectinload(Leg.departure_port),
            selectinload(PackingList.order).selectinload(Order.leg).selectinload(Leg.arrival_port),
            selectinload(PackingList.batches),
        ).where(PackingList.id == pl_id)
    )
    pl = result.scalar_one_or_none()
    if not pl:
        raise HTTPException(404)

    batch = next((b for b in pl.batches if b.id == batch_id), None)
    if not batch:
        raise HTTPException(404, "Batch non trouve")

    import shutil, zipfile, os

    template_path = "/app/app/static/BILL_OF_LADING_TEMPLATE.docx"
    if not os.path.exists(template_path):
        template_path = os.path.join(os.path.dirname(__file__), "..", "static", "BILL_OF_LADING_TEMPLATE.docx")
    if not os.path.exists(template_path):
        raise HTTPException(500, detail="Template BOL non trouve")

    voyage_id = pl.order.leg.leg_code if pl.order.leg else ""
    leg_id = pl.order.leg.id if pl.order.leg else None

    # BL number: use existing value if set, otherwise generate unique number for this leg
    bl_number = batch.bill_of_lading_id
    if not bl_number and leg_id:
        # Find all existing BL numbers for this leg to avoid duplicates
        existing_bls_result = await db.execute(
            select(PackingListBatch.bill_of_lading_id)
            .join(PackingList, PackingListBatch.packing_list_id == PackingList.id)
            .join(Order, PackingList.order_id == Order.id)
            .where(Order.leg_id == leg_id, PackingListBatch.bill_of_lading_id.isnot(None))
        )
        existing_bls = {r[0] for r in existing_bls_result.fetchall()}

        # Find next available sequence number
        seq = 1
        while True:
            candidate = f"TUAW_{voyage_id}_{seq:02d}"
            if candidate not in existing_bls:
                bl_number = candidate
                break
            seq += 1

        # Persist so it never changes
        batch.bill_of_lading_id = bl_number
        await db.flush()
    elif not bl_number:
        bl_number = f"TUAW_{voyage_id}_01"

    def _build_addr(b, prefix):
        if not b:
            return ""
        parts = filter(None, [
            getattr(b, f"{prefix}_name", None),
            getattr(b, f"{prefix}_address", None),
            " ".join(filter(None, [getattr(b, f"{prefix}_postal", None), getattr(b, f"{prefix}_city", None)])),
            getattr(b, f"{prefix}_country", None),
        ])
        return "\n".join(parts)

    packages_str = f"{batch.pallet_quantity or 0} palettes"
    if batch.cases_quantity:
        packages_str += f" / {batch.cases_quantity} cases"

    replacements = {
        "SHIPPER_NAME": _build_addr(batch, "shipper"),
        "BILL_OF_LADING_ID": bl_number,
        "BOOKING_CONFIRMATION_TOWT": pl.order.reference or "",
        "CONSIGNEE_ORDER_ADRESS": _build_addr(batch, "consignee"),
        "NOTIFY_ADRESS": _build_addr(batch, "notify"),
        "VESSEL": pl.order.leg.vessel.name if pl.order.leg and pl.order.leg.vessel else "",
        "VOYAGE_ID": voyage_id,
        "POL_NAME": pl.order.leg.departure_port.name if pl.order.leg and pl.order.leg.departure_port else "",
        "POD_NAME": pl.order.leg.arrival_port.name if pl.order.leg and pl.order.leg.arrival_port else "",
        "Maximum_de_CASES_QUANTITY": packages_str,
        "Nombre_de_PALLET_ID": str(batch.pallet_quantity or 0),
        "Maximum_de_WEIGHT_KG": str(int((batch.weight_kg or 0) * (batch.pallet_quantity or 1))),
        "TYPE_OF_GOODS": batch.description_of_goods or batch.type_of_goods or "",
        "DESCRIPTION_OF_GOODS": batch.description_of_goods or batch.type_of_goods or "",
        "PACKAGES": packages_str,
        "NUMBER_OF_OBL": "3",
    }

    work_dir = f"/tmp/bol_{pl.id}_{batch_id}_{datetime.now().strftime('%Y%m%d%H%M%S')}"
    os.makedirs(work_dir, exist_ok=True)
    output_path = os.path.join(work_dir, "output.docx")
    shutil.copy2(template_path, output_path)

    temp_docx = os.path.join(work_dir, "temp.docx")
    with zipfile.ZipFile(output_path, 'r') as zin:
        with zipfile.ZipFile(temp_docx, 'w') as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename == "word/document.xml":
                    content = data.decode('utf-8')
                    for key, val in replacements.items():
                        safe_val = (val or "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
                        safe_val = safe_val.replace("\n", "</w:t><w:br/><w:t>")
                        content = content.replace(f"\u00ab{key}\u00bb", safe_val)
                    data = content.encode('utf-8')
                zout.writestr(item, data)

    with open(temp_docx, 'rb') as f:
        buffer = io.BytesIO(f.read())
    shutil.rmtree(work_dir, ignore_errors=True)

    filename = f"{bl_number}_{datetime.now().strftime('%Y%m%d')}.docx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# === HTML BILL OF LADING VIEW ===
@router.get("/{pl_id}/bol/view", response_class=HTMLResponse)
async def bill_of_lading_view(
    pl_id: int, request: Request,
    user: User = Depends(require_permission("cargo", "C")),
    db: AsyncSession = Depends(get_db),
):
    """Render Bill of Lading as printable HTML page."""
    result = await db.execute(
        select(PackingList).options(
            selectinload(PackingList.order).selectinload(Order.leg).selectinload(Leg.vessel),
            selectinload(PackingList.order).selectinload(Order.leg).selectinload(Leg.departure_port),
            selectinload(PackingList.order).selectinload(Order.leg).selectinload(Leg.arrival_port),
            selectinload(PackingList.batches),
        ).where(PackingList.id == pl_id)
    )
    pl = result.scalar_one_or_none()
    if not pl:
        raise HTTPException(404)

    first = pl.batches[0] if pl.batches else None
    voyage_id = pl.order.leg.leg_code if pl.order.leg else ""
    leg_id = pl.order.leg.id if pl.order.leg else None
    bl_seq = 1
    if leg_id:
        all_pls_result = await db.execute(
            select(PackingList.id).join(Order).where(Order.leg_id == leg_id).order_by(PackingList.id)
        )
        all_pl_ids = [r[0] for r in all_pls_result.fetchall()]
        bl_seq = all_pl_ids.index(pl.id) + 1 if pl.id in all_pl_ids else 1
    bl_number = f"TUAW_{voyage_id}_{bl_seq:03d}"

    def _addr(batch, prefix):
        if not batch:
            return "—"
        parts = filter(None, [
            getattr(batch, f"{prefix}_name", None),
            getattr(batch, f"{prefix}_address", None),
            " ".join(filter(None, [getattr(batch, f"{prefix}_postal", None), getattr(batch, f"{prefix}_city", None)])),
            getattr(batch, f"{prefix}_country", None),
        ])
        return "\n".join(parts) or "—"

    return templates.TemplateResponse("cargo/bill_of_lading.html", {
        "request": request, "pl": pl, "bl_number": bl_number,
        "shipper_address": _addr(first, "shipper"),
        "consignee_address": _addr(first, "consignee"),
        "notify_address": _addr(first, "notify"),
        "now": datetime.now(),
    })


# === ARRIVAL NOTICE (PDF) ===
@router.get("/{pl_id}/arrival-notice")
async def arrival_notice(
    pl_id: int, request: Request,
    user: User = Depends(require_permission("cargo", "C")),
    db: AsyncSession = Depends(get_db),
):
    """Generate Arrival Notice document from packing list data."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    result = await db.execute(
        select(PackingList).options(
            selectinload(PackingList.order).selectinload(Order.leg).selectinload(Leg.vessel),
            selectinload(PackingList.order).selectinload(Order.leg).selectinload(Leg.departure_port),
            selectinload(PackingList.order).selectinload(Order.leg).selectinload(Leg.arrival_port),
            selectinload(PackingList.batches),
        ).where(PackingList.id == pl_id)
    )
    pl = result.scalar_one_or_none()
    if not pl:
        raise HTTPException(404)

    leg = pl.order.leg if pl.order else None
    if not leg:
        raise HTTPException(404, detail="Aucun leg associé à cette commande")
    first = pl.batches[0] if pl.batches else None
    total_pallets = sum(b.pallet_quantity or 0 for b in pl.batches)
    total_weight = sum(b.weight_kg or 0 for b in pl.batches)
    total_volume = sum(b.volume_m3 or 0 for b in pl.batches)
    goods_types = ", ".join(set(b.type_of_goods for b in pl.batches if b.type_of_goods))

    # Build BL number
    voyage_id = leg.leg_code if leg else ""
    leg_id_val = leg.id if leg else None
    bl_seq = 1
    if leg_id_val:
        all_pls_result = await db.execute(
            select(PackingList.id).join(Order).where(Order.leg_id == leg_id_val).order_by(PackingList.id)
        )
        all_pl_ids = [r[0] for r in all_pls_result.fetchall()]
        bl_seq = all_pl_ids.index(pl.id) + 1 if pl.id in all_pl_ids else 1
    bl_number = f"TUAW_{voyage_id}_{bl_seq:03d}"

    def _addr(batch, prefix):
        if not batch:
            return "—"
        parts = filter(None, [
            getattr(batch, f"{prefix}_name", None),
            getattr(batch, f"{prefix}_address", None),
            " ".join(filter(None, [getattr(batch, f"{prefix}_postal", None), getattr(batch, f"{prefix}_city", None)])),
            getattr(batch, f"{prefix}_country", None),
        ])
        return "\n".join(parts) or "—"

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=20*mm, bottomMargin=15*mm, leftMargin=15*mm, rightMargin=15*mm)
    styles = getSampleStyleSheet()
    hdr_color = colors.HexColor("#095561")
    elements = []

    # Title
    title_style = ParagraphStyle("ANTitle", parent=styles["Heading1"], fontSize=18, textColor=hdr_color, spaceAfter=4)
    elements.append(Paragraph("ARRIVAL NOTICE", title_style))

    sub_style = ParagraphStyle("Sub", parent=styles["Normal"], fontSize=10, textColor=colors.HexColor("#555"), spaceAfter=2)
    elements.append(Paragraph("TOWT — Transport à la Voile", sub_style))
    elements.append(Spacer(1, 8*mm))

    # Header info table
    info_data = [
        ["B/L Number", bl_number, "Booking Ref.", pl.order.reference or "—"],
        ["Vessel", leg.vessel.name if leg and leg.vessel else "—", "Voyage", voyage_id or "—"],
        ["Port of Loading", (leg.departure_port.name if leg and leg.departure_port else "—"), "Port of Discharge", (leg.arrival_port.name if leg and leg.arrival_port else "—")],
        ["ETA", leg.eta.strftime("%d/%m/%Y %H:%M") if leg and leg.eta else "TBC", "ETD", leg.etd.strftime("%d/%m/%Y %H:%M") if leg and leg.etd else "TBC"],
    ]
    info_t = Table(info_data, colWidths=[100, 170, 100, 170])
    info_t.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#ddd")),
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#f0f7f8")),
        ("BACKGROUND", (2, 0), (2, -1), colors.HexColor("#f0f7f8")),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    elements.append(info_t)
    elements.append(Spacer(1, 6*mm))

    # Parties
    parties_data = [
        ["Consignee", _addr(first, "consignee")],
        ["Notify Party", _addr(first, "notify")],
        ["Shipper", _addr(first, "shipper")],
    ]
    parties_t = Table(parties_data, colWidths=[100, 440])
    parties_t.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#ddd")),
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#f0f7f8")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    elements.append(parties_t)
    elements.append(Spacer(1, 6*mm))

    # Cargo summary
    cargo_hdr = ParagraphStyle("CH", parent=styles["Heading2"], fontSize=12, textColor=hdr_color, spaceAfter=4)
    elements.append(Paragraph("Cargo Summary", cargo_hdr))

    cargo_data = [["Description", "Pallets", "Packages", "Gross Weight (kg)", "Volume (m³)"]]
    for b in pl.batches:
        cargo_data.append([
            b.description_of_goods or b.type_of_goods or "—",
            str(b.pallet_quantity or "—"),
            str(b.cases_quantity or "—"),
            str(b.weight_kg or "—"),
            str(b.volume_m3 or "—"),
        ])
    cargo_data.append(["TOTAL", str(total_pallets), "", f"{int(total_weight)}", f"{total_volume:.2f}" if total_volume else "—"])

    cargo_t = Table(cargo_data, colWidths=[200, 60, 60, 100, 80])
    cargo_t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), hdr_color),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#ddd")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#f9fafb")]),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#f0f7f8")),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    elements.append(cargo_t)
    elements.append(Spacer(1, 10*mm))

    # Footer
    footer_style = ParagraphStyle("Footer", parent=styles["Normal"], fontSize=8, textColor=colors.HexColor("#aaa"))
    elements.append(Paragraph(
        f"Generated on {datetime.now().strftime('%d/%m/%Y %H:%M')} — TOWT Operations Platform<br/>"
        f"Number of Original Bills of Lading: THREE (3)",
        footer_style,
    ))

    doc.build(elements)
    buf.seek(0)

    fname = f"ArrivalNotice_{bl_number}_{datetime.now().strftime('%Y%m%d')}.pdf"
    return StreamingResponse(
        buf, media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


# === ADD BATCH (exploitation) ===
@router.post("/{pl_id}/add-batch", response_class=HTMLResponse)
async def add_batch(
    pl_id: int, request: Request,
    user: User = Depends(require_permission("cargo", "M")),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(
        select(PackingList).options(
            selectinload(PackingList.batches),
            selectinload(PackingList.order).selectinload(Order.leg).selectinload(Leg.vessel),
            selectinload(PackingList.order).selectinload(Order.leg).selectinload(Leg.departure_port),
            selectinload(PackingList.order).selectinload(Order.leg).selectinload(Leg.arrival_port),
        ).where(PackingList.id == pl_id)
    )
    pl = result.scalar_one_or_none()
    if not pl or pl.is_locked:
        raise HTTPException(403)

    first = pl.batches[0] if pl.batches else None
    batch = PackingListBatch(
        packing_list_id=pl.id,
        batch_number=len(pl.batches) + 1,
        voyage_id=first.voyage_id if first else None,
        vessel=first.vessel if first else None,
        loading_date=first.loading_date if first else None,
        pol_code=first.pol_code if first else None,
        pod_code=first.pod_code if first else None,
        pol_name=first.pol_name if first else None,
        pod_name=first.pod_name if first else None,
        booking_confirmation=first.booking_confirmation if first else None,
        freight_rate=first.freight_rate if first else None,
    )
    db.add(batch)
    await db.flush()

    url = f"/cargo/{pl_id}"
    if request.headers.get("HX-Request"):
        return HTMLResponse(content="", headers={"HX-Redirect": url})
    return RedirectResponse(url=url, status_code=303)


# === AUDIT HELPER ===
CLIENT_FIELDS = [
    'customer_name', 'freight_forwarder', 'code_transitaire', 'shipper_name',
    'shipper_address', 'po_number', 'customer_batch_id', 'notify_address', 'consignee_address',
    'pallet_type', 'type_of_goods', 'bio_products', 'cases_quantity',
    'units_per_case', 'imo_product_class', 'pallet_quantity',
    'length_cm', 'width_cm', 'height_cm', 'weight_kg', 'cargo_value_usd',
]

async def audit_batch_changes(db: AsyncSession, pl_id: int, batch: PackingListBatch, form_data: dict, changed_by: str):
    """Record all field changes for a batch."""
    for field in CLIENT_FIELDS:
        old_val = getattr(batch, field, None)
        new_val = form_data.get(field)
        if new_val is not None:
            old_str = str(old_val) if old_val is not None else ""
            new_str = str(new_val).strip()
            if old_str != new_str and (old_str or new_str):
                db.add(PackingListAudit(
                    packing_list_id=pl_id,
                    batch_id=batch.id,
                    field_name=field,
                    old_value=old_str or None,
                    new_value=new_str or None,
                    changed_by=changed_by,
                ))


def apply_batch_fields(batch, form_data):
    """Apply client-editable fields to a batch from form data."""
    batch.customer_name = form_data.get('customer_name', batch.customer_name)
    batch.freight_forwarder = form_data.get('freight_forwarder', batch.freight_forwarder)
    batch.code_transitaire = form_data.get('code_transitaire', batch.code_transitaire)
    batch.shipper_name = form_data.get('shipper_name', batch.shipper_name)
    batch.shipper_address = form_data.get('shipper_address', batch.shipper_address)
    batch.po_number = form_data.get('po_number', batch.po_number)
    batch.customer_batch_id = form_data.get('customer_batch_id', batch.customer_batch_id)
    batch.notify_address = form_data.get('notify_address', batch.notify_address)
    batch.consignee_address = form_data.get('consignee_address', batch.consignee_address)
    batch.pallet_type = form_data.get('pallet_type', batch.pallet_type)
    batch.type_of_goods = form_data.get('type_of_goods', batch.type_of_goods)
    batch.bio_products = form_data.get('bio_products', batch.bio_products)
    batch.cases_quantity = pi(form_data.get('cases_quantity'))
    batch.units_per_case = pi(form_data.get('units_per_case'))
    batch.imo_product_class = form_data.get('imo_product_class', batch.imo_product_class)
    batch.un_number = form_data.get('un_number', batch.un_number)
    batch.pallet_quantity = pi(form_data.get('pallet_quantity'))
    batch.length_cm = pf(form_data.get('length_cm'))
    batch.width_cm = pf(form_data.get('width_cm'))
    batch.height_cm = pf(form_data.get('height_cm'))
    batch.weight_kg = pf(form_data.get('weight_kg'))
    batch.cargo_value_usd = pf(form_data.get('cargo_value_usd'))
    batch.compute_dimensions()


# IMDG validation — returns list of warnings for dangerous goods
IMDG_REQUIRED_FIELDS = {
    "un_number": "UN Number",
    "type_of_goods": "Description des marchandises",
    "weight_kg": "Poids",
}


def validate_imdg(batch) -> list[str]:
    """Validate IMDG requirements for dangerous goods batches."""
    warnings = []
    if batch.imo_product_class and batch.imo_product_class != "Non-Dangerous Goods":
        for field, label in IMDG_REQUIRED_FIELDS.items():
            val = getattr(batch, field, None)
            if val is None or (isinstance(val, str) and not val.strip()):
                warnings.append(f"Batch #{batch.batch_number}: {label} requis pour marchandises dangereuses ({batch.imo_product_class})")
    return warnings


# === EXPLOITATION EDIT BATCHES (with audit) ===
@router.post("/{pl_id}/edit", response_class=HTMLResponse)
async def exploitation_edit_batches(
    pl_id: int, request: Request,
    user: User = Depends(require_permission("cargo", "M")),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(
        select(PackingList).options(selectinload(PackingList.batches))
        .where(PackingList.id == pl_id)
    )
    pl = result.scalar_one_or_none()
    if not pl:
        raise HTTPException(404)

    form = await request.form()
    changed_by = user.full_name or user.username

    for batch in pl.batches:
        prefix = f"batch_{batch.id}_"
        form_data = {k.replace(prefix, ''): v for k, v in form.items() if k.startswith(prefix)}
        if form_data:
            await audit_batch_changes(db, pl.id, batch, form_data, changed_by)
            apply_batch_fields(batch, form_data)

    await db.flush()
    url = f"/cargo/{pl_id}"
    if request.headers.get("HX-Request"):
        return HTMLResponse(content="", headers={"HX-Redirect": url})
    return RedirectResponse(url=url, status_code=303)


# === AUDIT LOG VIEW ===
@router.get("/{pl_id}/history", response_class=HTMLResponse)
async def audit_history(
    pl_id: int, request: Request,
    user: User = Depends(require_permission("cargo", "C")),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(
        select(PackingList).options(selectinload(PackingList.order))
        .where(PackingList.id == pl_id)
    )
    pl = result.scalar_one_or_none()
    if not pl:
        raise HTTPException(404)

    audit_result = await db.execute(
        select(PackingListAudit).where(PackingListAudit.packing_list_id == pl_id)
        .order_by(PackingListAudit.changed_at.desc())
    )
    logs = audit_result.scalars().all()

    return templates.TemplateResponse("cargo/history.html", {
        "request": request, "user": user,
        "pl": pl, "logs": logs,
        "active_module": "cargo",
    })


# === EXCEL IMPORT ===
@router.post("/{pl_id}/import-excel", response_class=HTMLResponse)
async def import_excel(
    pl_id: int, request: Request,
    user: User = Depends(require_permission("cargo", "M")),
    db: AsyncSession = Depends(get_db),
):
    from fastapi import UploadFile, File
    result = await db.execute(
        select(PackingList).options(
            selectinload(PackingList.batches),
            selectinload(PackingList.order).selectinload(Order.leg).selectinload(Leg.vessel),
            selectinload(PackingList.order).selectinload(Order.leg).selectinload(Leg.departure_port),
            selectinload(PackingList.order).selectinload(Order.leg).selectinload(Leg.arrival_port),
        ).where(PackingList.id == pl_id)
    )
    pl = result.scalar_one_or_none()
    if not pl or pl.is_locked:
        raise HTTPException(403)

    form = await request.form()
    file = form.get("file")
    if not file:
        raise HTTPException(400, detail="Aucun fichier")

    import openpyxl
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active

    # Read headers from row 1
    headers = [cell.value for cell in ws[1] if cell.value]
    header_map = {h: idx for idx, h in enumerate(headers)}

    changed_by = user.full_name or user.username

    # Map Excel columns to model fields
    EXCEL_TO_FIELD = {
        "CUSTOMER_NAME": "customer_name",
        "FREIGHT_FORWARDER": "freight_forwarder",
        "CODE_TRANSITAIRE": "code_transitaire",
        "SHIPPER_NAME": "shipper_name",
        "PO_NUMBER": "po_number",
        "CUSTOMER_BATCH_ID": "customer_batch_id",
        "NOTIFY_ADRESS": "notify_address",
        "CONSIGNEE_ORDER_ADRESS": "consignee_address",
        "PALLET_TYPE": "pallet_type",
        "TYPE_OF_GOODS": "type_of_goods",
        "BIO_PRODUCTS": "bio_products",
        "CASES_QUANTITY": "cases_quantity",
        "UNITS_PER_CASE": "units_per_case",
        "IMO_PRODUCT_CLASS": "imo_product_class",
        "PALLET_QUANTITY_PER_BATCH": "pallet_quantity",
        "LENGTH_CM": "length_cm",
        "WIDTH_CM": "width_cm",
        "HEIGHT_CM": "height_cm",
        "WEIGHT_KG": "weight_kg",
        "CARGO_VALUE_USD": "cargo_value_usd",
    }

    # Delete existing batches and recreate from Excel
    for b in pl.batches:
        await db.delete(b)
    await db.flush()

    first_batch = None
    batch_num = 0
    for row in ws.iter_rows(min_row=2, values_only=False):
        values = [cell.value for cell in row]
        if not any(values):
            continue
        batch_num += 1
        batch = PackingListBatch(packing_list_id=pl.id, batch_number=batch_num)

        # Pre-fill TOWT fields from order/leg
        batch.booking_confirmation = pl.order.reference
        batch.customer_name = pl.order.client_name
        batch.freight_rate = pl.order.unit_price
        if pl.order.leg:
            leg = pl.order.leg
            batch.voyage_id = leg.leg_code
            batch.vessel = leg.vessel.name if leg.vessel else None
            batch.loading_date = leg.etd.date() if leg.etd else None
            batch.pol_code = leg.departure_port.locode if leg.departure_port else None
            batch.pod_code = leg.arrival_port.locode if leg.arrival_port else None
            batch.pol_name = leg.departure_port.name if leg.departure_port else None
            batch.pod_name = leg.arrival_port.name if leg.arrival_port else None

        # Apply Excel data
        for excel_col, model_field in EXCEL_TO_FIELD.items():
            if excel_col in header_map:
                idx = header_map[excel_col]
                if idx < len(values):
                    val = values[idx]
                    if val is not None:
                        setattr(batch, model_field, val)

        batch.compute_dimensions()
        db.add(batch)

        # Audit log
        db.add(PackingListAudit(
            packing_list_id=pl.id, batch_id=None,
            field_name="excel_import",
            old_value=None,
            new_value=f"Batch #{batch_num} importé depuis Excel",
            changed_by=changed_by,
        ))

    await db.flush()
    url = f"/cargo/{pl_id}"
    if request.headers.get("HX-Request"):
        return HTMLResponse(content="", headers={"HX-Redirect": url})
    return RedirectResponse(url=url, status_code=303)


# === VOYAGE EXPORT (all packing lists for a leg) ===
@router.get("/voyage/{leg_id}/excel")
async def export_voyage_excel(
    leg_id: int,
    user: User = Depends(require_permission("cargo", "C")),
    db: AsyncSession = Depends(get_db),
):
    leg_result = await db.execute(
        select(Leg).options(selectinload(Leg.vessel), selectinload(Leg.departure_port), selectinload(Leg.arrival_port))
        .where(Leg.id == leg_id)
    )
    leg = leg_result.scalar_one_or_none()
    if not leg:
        raise HTTPException(404)

    # Find all orders assigned to this leg, then their packing lists
    orders_result = await db.execute(
        select(Order).where(Order.leg_id == leg_id)
    )
    orders = orders_result.scalars().all()
    order_ids = [o.id for o in orders]

    if not order_ids:
        raise HTTPException(404, detail="Aucune commande sur ce voyage")

    pls_result = await db.execute(
        select(PackingList).options(selectinload(PackingList.batches), selectinload(PackingList.order))
        .where(PackingList.order_id.in_(order_ids))
    )
    packing_lists = pls_result.scalars().all()

    all_batches = []
    for pl in packing_lists:
        all_batches.extend(pl.batches)

    if not all_batches:
        raise HTTPException(404, detail="Aucun batch")

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "LOADING_DATA"

    headers = [
        "VOYAGE_ID", "VESSEL", "LOADING_DATE", "POL_CODE", "POD_CODE",
        "POL_NAME", "POD_NAME", "CUSTOMER_NAME", "BOOKING_CONFIRMATION_TOWT",
        "FREIGHT_RATE_PER_PALLET_EURO", "CARGO_VALUE_USD", "FREIGHT_FORWARDER",
        "CODE_TRANSITAIRE", "SHIPPER_NAME", "PO_NUMBER", "CUSTOMER_BATCH_ID",
        "NOTIFY_ADRESS", "CONSIGNEE_ORDER_ADRESS", "PALLET_TYPE",
        "TYPE_OF_GOODS", "BIO_PRODUCTS", "CASES_QUANTITY", "UNITS_PER_CASE",
        "IMO_PRODUCT_CLASS", "PALLET_QUANTITY_PER_BATCH",
        "LENGTH_CM", "WIDTH_CM", "HEIGHT_CM", "WEIGHT_KG", "SURFACE_M2", "VOLUME_M3",
    ]

    header_fill = PatternFill(start_color="095561", end_color="095561", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = max(15, len(h) + 2)

    for row_idx, batch in enumerate(all_batches, 2):
        data = [
            batch.voyage_id, batch.vessel,
            batch.loading_date.isoformat() if batch.loading_date else "",
            batch.pol_code, batch.pod_code, batch.pol_name, batch.pod_name,
            batch.customer_name, batch.booking_confirmation, batch.freight_rate,
            batch.cargo_value_usd, batch.freight_forwarder, batch.code_transitaire,
            batch.shipper_name, batch.po_number, batch.customer_batch_id,
            batch.notify_address, batch.consignee_address, batch.pallet_type,
            batch.type_of_goods, batch.bio_products, batch.cases_quantity,
            batch.units_per_case, batch.imo_product_class, batch.pallet_quantity,
            batch.length_cm, batch.width_cm, batch.height_cm, batch.weight_kg,
            batch.surface_m2, batch.volume_m3,
        ]
        for col_idx, val in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val if val is not None else "")
            cell.border = thin_border

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"PACKING_LIST_VOYAGE_{leg.leg_code}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ═══════════════════════════════════════════════════════════════
# EXTERNAL CLIENT ACCESS (no auth required)
# ═══════════════════════════════════════════════════════════════

# ═══ BACKOFFICE CARGO MESSAGING ═══════════════════════════════

@router.post("/{pl_id}/messages/send", response_class=HTMLResponse)
async def backoffice_cargo_send_message(
    pl_id: int, request: Request,
    user: User = Depends(require_permission("cargo", "M")),
    db: AsyncSession = Depends(get_db),
):
    form = await request.form()
    message_text = form.get("message", "").strip()
    if message_text:
        msg = PortalMessage(
            packing_list_id=pl_id,
            sender_type="company",
            sender_name=user.username or "TOWT",
            message=message_text,
        )
        db.add(msg)
        await db.flush()
    return RedirectResponse(url=f"/cargo/{pl_id}#messaging", status_code=303)


@router.post("/{pl_id}/messages/read", response_class=HTMLResponse)
async def backoffice_cargo_mark_messages_read(
    pl_id: int, request: Request,
    user: User = Depends(require_permission("cargo", "M")),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(
        select(PortalMessage).where(
            PortalMessage.packing_list_id == pl_id,
            PortalMessage.sender_type == "client",
            PortalMessage.is_read == False,
        )
    )
    for msg in result.scalars().all():
        msg.is_read = True
    await db.flush()
    return RedirectResponse(url=f"/cargo/{pl_id}#messaging", status_code=303)


# ═══ CLIENT EXTERNAL PORTAL ═══════════════════════════════════

ext_router = APIRouter(prefix="/p", tags=["packing-external"])

# ── Helpers ────────────────────────────────────────────────────
VALID_LANGS = ('fr', 'en', 'es', 'pt-br', 'vi')

def _lang(request):
    lang = request.query_params.get('lang') or request.cookies.get('towt_lang') or 'fr'
    return lang if lang in VALID_LANGS else 'fr'

async def _get_pl(token, db, request=None):
    from datetime import datetime, timezone as tz
    from app.utils.portal_security import check_token_rate_limit, record_token_attempt, log_portal_access
    if request:
        await check_token_rate_limit(request, db)
    result = await db.execute(
        select(PackingList).options(
            selectinload(PackingList.order).selectinload(Order.leg).selectinload(Leg.vessel),
            selectinload(PackingList.order).selectinload(Order.leg).selectinload(Leg.departure_port),
            selectinload(PackingList.order).selectinload(Order.leg).selectinload(Leg.arrival_port),
            selectinload(PackingList.batches),
        ).where(PackingList.token == token)
    )
    pl = result.scalar_one_or_none()
    if not pl:
        if request:
            await record_token_attempt(request, db)
            await db.commit()
        raise HTTPException(404, detail="Lien invalide ou expiré")
    # Check token expiration
    if pl.token_expires_at and pl.token_expires_at < datetime.now(tz.utc):
        raise HTTPException(404, detail="Lien invalide ou expiré")
    # Audit trail
    if request:
        await log_portal_access(db, request, "cargo", token, packing_list_id=pl.id)
    return pl

async def _unread_count(pl_id, db):
    r = await db.execute(
        select(func.count(PortalMessage.id)).where(
            PortalMessage.packing_list_id == pl_id,
            PortalMessage.sender_type == "company",
            PortalMessage.is_read == False,
        )
    )
    return r.scalar() or 0


# ── Default: redirect to Packing List (primary client action) ─
@ext_router.get("/{token}", response_class=HTMLResponse)
async def client_portal_default(token: str, request: Request, db: AsyncSession = Depends(get_db)):
    pl = await _get_pl(token, db, request)
    lang = _lang(request)
    return RedirectResponse(url=f"/p/{token}/packing?lang={lang}", status_code=303)


# ── Privacy policy ────────────────────────────────────────────
@ext_router.get("/{token}/privacy", response_class=HTMLResponse)
async def client_portal_privacy(token: str, request: Request, db: AsyncSession = Depends(get_db)):
    pl = await _get_pl(token, db, request)
    lang = _lang(request)
    return templates.TemplateResponse("cargo/portal_privacy.html", {
        "request": request, "pl": pl,
        "lang": lang, "active_page": "privacy", "page_suffix": "/privacy",
        "unread_messages": await _unread_count(pl.id, db),
    })


# ── Page 1: Packing List ─────────────────────────────────────
@ext_router.get("/{token}/packing", response_class=HTMLResponse)
async def client_portal_packing(token: str, request: Request, db: AsyncSession = Depends(get_db)):
    pl = await _get_pl(token, db, request)
    lang = _lang(request)
    return templates.TemplateResponse("cargo/portal_packing.html", {
        "request": request, "pl": pl, "imo_classes": IMO_CLASSES,
        "lang": lang, "active_page": "packing", "page_suffix": "/packing",
        "unread_messages": await _unread_count(pl.id, db),
    })


# ── Page 2: Vessel ────────────────────────────────────────────
@ext_router.get("/{token}/vessel", response_class=HTMLResponse)
async def client_portal_vessel(token: str, request: Request, db: AsyncSession = Depends(get_db)):
    pl = await _get_pl(token, db, request)
    lang = _lang(request)
    vessel = pl.order.leg.vessel if pl.order.leg else None
    return templates.TemplateResponse("cargo/portal_vessel.html", {
        "request": request, "pl": pl, "vessel": vessel,
        "lang": lang, "active_page": "vessel", "page_suffix": "/vessel",
        "unread_messages": await _unread_count(pl.id, db),
    })


# ── Page 3: Voyage ────────────────────────────────────────────
@ext_router.get("/{token}/voyage", response_class=HTMLResponse)
async def client_portal_voyage(token: str, request: Request, db: AsyncSession = Depends(get_db)):
    pl = await _get_pl(token, db, request)
    lang = _lang(request)
    leg = pl.order.leg
    itinerary = []
    crew = []
    if leg:
        # Full itinerary for same vessel/year
        yr = leg.etd.year if leg.etd else None
        if yr:
            it_result = await db.execute(
                select(Leg).options(
                    selectinload(Leg.departure_port), selectinload(Leg.arrival_port)
                ).where(Leg.vessel_id == leg.vessel_id)
                .order_by(Leg.etd)
            )
            itinerary = it_result.scalars().all()
        # Crew on board
        from datetime import date
        crew_result = await db.execute(
            select(CrewAssignment).options(selectinload(CrewAssignment.member))
            .where(
                CrewAssignment.vessel_id == leg.vessel_id,
                CrewAssignment.embark_date <= (leg.etd.date() if leg.etd else date.today()),
                (CrewAssignment.disembark_date == None) | (CrewAssignment.disembark_date >= (leg.etd.date() if leg.etd else date.today())),
            )
        )
        crew = crew_result.scalars().all()
    return templates.TemplateResponse("cargo/portal_voyage.html", {
        "request": request, "pl": pl, "leg": leg,
        "itinerary": itinerary, "crew": crew,
        "lang": lang, "active_page": "voyage", "page_suffix": "/voyage",
        "unread_messages": await _unread_count(pl.id, db),
    })


# ── Page 3b: Stowage Position ─────────────────────────────────
@ext_router.get("/{token}/stowage", response_class=HTMLResponse)
async def client_portal_stowage(token: str, request: Request, db: AsyncSession = Depends(get_db)):
    pl = await _get_pl(token, db, request)
    lang = _lang(request)
    leg = pl.order.leg

    from app.models.stowage import (
        StowagePlan, ZONE_DEFINITIONS, LOADING_ORDER, DANGEROUS_ZONES, get_zone_label
    )
    batch_positions = []
    if leg:
        for batch in pl.batches:
            result = await db.execute(
                select(StowagePlan).where(
                    StowagePlan.leg_id == leg.id,
                    StowagePlan.batch_id == batch.id,
                )
            )
            plan = result.scalar_one_or_none()
            batch_positions.append({
                "batch": batch,
                "zone_code": plan.zone_code if plan else None,
                "zone_label": get_zone_label(plan.zone_code, lang) if plan else None,
            })

    return templates.TemplateResponse("cargo/portal_stowage.html", {
        "request": request, "pl": pl,
        "batch_positions": batch_positions,
        "zone_definitions": ZONE_DEFINITIONS,
        "loading_order": LOADING_ORDER,
        "dangerous_zones": DANGEROUS_ZONES,
        "lang": lang, "active_page": "stowage", "page_suffix": "/stowage",
        "unread_messages": await _unread_count(pl.id, db),
    })


# ── Page: Guide ──────────────────────────────────────────────
@ext_router.get("/{token}/guide", response_class=HTMLResponse)
async def client_portal_guide(token: str, request: Request, db: AsyncSession = Depends(get_db)):
    pl = await _get_pl(token, db, request)
    lang = _lang(request)
    return templates.TemplateResponse("cargo/client_guide.html", {
        "request": request, "pl": pl, "token": token,
        "lang": lang, "active_page": "guide", "page_suffix": "/guide",
        "unread_messages": await _unread_count(pl.id, db),
    })


# ── Page 4: Documents ─────────────────────────────────────────
CARGO_UPLOAD_DIR = "/app/uploads/cargo_docs"

@ext_router.get("/{token}/documents", response_class=HTMLResponse)
async def client_portal_docs(token: str, request: Request, db: AsyncSession = Depends(get_db)):
    pl = await _get_pl(token, db, request)
    lang = _lang(request)
    docs_result = await db.execute(
        select(PackingListDocument)
        .where(PackingListDocument.packing_list_id == pl.id)
        .order_by(PackingListDocument.created_at.desc())
    )
    uploaded_docs = docs_result.scalars().all()
    return templates.TemplateResponse("cargo/portal_docs.html", {
        "request": request, "pl": pl,
        "uploaded_docs": uploaded_docs,
        "lang": lang, "active_page": "docs", "page_suffix": "/documents",
        "unread_messages": await _unread_count(pl.id, db),
    })


@ext_router.post("/{token}/documents/upload", response_class=HTMLResponse)
async def client_upload_doc(
    token: str, request: Request,
    file: UploadFile = File(...),
    notes: str = Form(""),
    db: AsyncSession = Depends(get_db),
):
    import os, uuid
    pl = await _get_pl(token, db, request)
    lang = _lang(request)

    if not file.filename:
        raise HTTPException(400, "Fichier manquant")
    content = await file.read()
    if len(content) > 20 * 1024 * 1024:
        raise HTTPException(400, "Fichier trop volumineux (max 20 Mo)")

    os.makedirs(CARGO_UPLOAD_DIR, exist_ok=True)
    ext = os.path.splitext(file.filename)[1]
    safe_name = f"{pl.id}_{uuid.uuid4().hex[:8]}{ext}"
    file_path = os.path.join(CARGO_UPLOAD_DIR, safe_name)
    with open(file_path, "wb") as f:
        f.write(content)

    doc = PackingListDocument(
        packing_list_id=pl.id,
        filename=file.filename,
        file_path=file_path,
        file_size=len(content),
        uploaded_by="Client",
        notes=notes.strip() or None,
    )
    db.add(doc)
    await db.flush()

    from app.models.notification import Notification
    db.add(Notification(
        type="cargo_document_uploaded",
        title=f"Document depose — {pl.order.reference}",
        detail=file.filename,
        link=f"/cargo/{pl.id}",
    ))
    await db.flush()

    return RedirectResponse(url=f"/p/{token}/documents?lang={lang}&uploaded=1", status_code=303)


@ext_router.get("/{token}/documents/{doc_id}/download")
async def client_download_doc(token: str, doc_id: int, db: AsyncSession = Depends(get_db)):
    pl = await _get_pl(token, db)
    doc = await db.get(PackingListDocument, doc_id)
    if not doc or doc.packing_list_id != pl.id:
        raise HTTPException(404)
    import os
    if not os.path.exists(doc.file_path):
        raise HTTPException(404, "Fichier introuvable")
    return FileResponse(doc.file_path, filename=doc.filename)


@ext_router.post("/{token}/documents/{doc_id}/delete", response_class=HTMLResponse)
async def client_delete_doc(token: str, doc_id: int, request: Request, db: AsyncSession = Depends(get_db)):
    pl = await _get_pl(token, db, request)
    lang = _lang(request)
    doc = await db.get(PackingListDocument, doc_id)
    if not doc or doc.packing_list_id != pl.id:
        raise HTTPException(404)
    if doc.uploaded_by != "Client":
        raise HTTPException(403, "Vous ne pouvez supprimer que vos propres documents")
    import os
    if os.path.exists(doc.file_path):
        os.remove(doc.file_path)
    await db.delete(doc)
    await db.flush()
    return RedirectResponse(url=f"/p/{token}/documents?lang={lang}", status_code=303)


# ── Page 5: Messages ──────────────────────────────────────────
@ext_router.get("/{token}/messages", response_class=HTMLResponse)
async def client_portal_messages(token: str, request: Request, db: AsyncSession = Depends(get_db)):
    pl = await _get_pl(token, db, request)
    lang = _lang(request)
    # Load messages
    msg_result = await db.execute(
        select(PortalMessage)
        .where(PortalMessage.packing_list_id == pl.id)
        .order_by(PortalMessage.created_at)
    )
    messages = msg_result.scalars().all()
    # Mark company messages as read
    for m in messages:
        if m.sender_type == "company" and not m.is_read:
            m.is_read = True
    await db.flush()
    return templates.TemplateResponse("cargo/portal_messages.html", {
        "request": request, "pl": pl, "messages": messages,
        "lang": lang, "active_page": "messages", "page_suffix": "/messages",
        "unread_messages": 0,
    })


@ext_router.post("/{token}/messages/send", response_class=HTMLResponse)
async def client_portal_send_message(token: str, request: Request, db: AsyncSession = Depends(get_db)):
    pl = await _get_pl(token, db, request)
    lang = _lang(request)
    form = await request.form()
    message_text = form.get("message", "").strip()
    if message_text:
        msg = PortalMessage(
            packing_list_id=pl.id,
            sender_type="client",
            sender_name=pl.order.client_name or "Client",
            message=message_text,
        )
        db.add(msg)
        # Notification
        from app.models.notification import Notification
        db.add(Notification(
            type="new_cargo_message",
            title=f"Message client — {pl.order.reference}",
            detail=f"{pl.order.client_name}: {message_text[:100]}",
            link=f"/cargo/{pl.id}#messages",
            packing_list_id=pl.id,
        ))
        await db.flush()
    return RedirectResponse(url=f"/p/{token}/messages?lang={lang}", status_code=303)


# ── Existing actions (batch add/save/delete) ──────────────────
@ext_router.post("/{token}/batch/add", response_class=HTMLResponse)
async def client_add_batch(token: str, request: Request, db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(PackingList).options(
            selectinload(PackingList.order).selectinload(Order.leg).selectinload(Leg.vessel),
            selectinload(PackingList.order).selectinload(Order.leg).selectinload(Leg.departure_port),
            selectinload(PackingList.order).selectinload(Order.leg).selectinload(Leg.arrival_port),
            selectinload(PackingList.batches),
        ).where(PackingList.token == token)
    )
    pl = result.scalar_one_or_none()
    if not pl or pl.is_locked:
        raise HTTPException(403)
    first = pl.batches[0] if pl.batches else None
    batch = PackingListBatch(
        packing_list_id=pl.id,
        batch_number=len(pl.batches) + 1,
        voyage_id=first.voyage_id if first else None,
        vessel=first.vessel if first else None,
        loading_date=first.loading_date if first else None,
        pol_code=first.pol_code if first else None,
        pod_code=first.pod_code if first else None,
        pol_name=first.pol_name if first else None,
        pod_name=first.pod_name if first else None,
        booking_confirmation=first.booking_confirmation if first else None,
        freight_rate=first.freight_rate if first else None,
    )
    db.add(batch)
    await db.flush()
    return RedirectResponse(url=f"/p/{token}/packing", status_code=303)


@ext_router.post("/{token}/save", response_class=HTMLResponse)
async def client_save_batches(token: str, request: Request, db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(PackingList).options(selectinload(PackingList.batches))
        .where(PackingList.token == token)
    )
    pl = result.scalar_one_or_none()
    if not pl or pl.is_locked:
        raise HTTPException(403)
    form = await request.form()
    for batch in pl.batches:
        prefix = f"batch_{batch.id}_"
        form_data = {k.replace(prefix, ''): v for k, v in form.items() if k.startswith(prefix)}
        if form_data:
            await audit_batch_changes(db, pl.id, batch, form_data, "Client")
            apply_batch_fields(batch, form_data)
    if pl.status == "draft":
        pl.status = "submitted"
    await db.flush()
    return RedirectResponse(url=f"/p/{token}/packing?saved=1", status_code=303)


@ext_router.delete("/{token}/batch/{batch_id}", response_class=HTMLResponse)
async def client_delete_batch(token: str, batch_id: int, request: Request, db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(PackingList).options(selectinload(PackingList.batches))
        .where(PackingList.token == token)
    )
    pl = result.scalar_one_or_none()
    if not pl or pl.is_locked:
        raise HTTPException(403)
    if len(pl.batches) <= 1:
        raise HTTPException(400, detail="Au moins 1 batch requis")
    batch_result = await db.execute(
        select(PackingListBatch).where(
            PackingListBatch.id == batch_id,
            PackingListBatch.packing_list_id == pl.id,
        )
    )
    batch = batch_result.scalar_one_or_none()
    if batch:
        await db.delete(batch)
        await db.flush()
    return RedirectResponse(url=f"/p/{token}/packing", status_code=303)


# ── Excel Template Download (external portal) ────────────────
@ext_router.get("/{token}/template")
async def client_download_template(token: str, request: Request, db: AsyncSession = Depends(get_db)):
    """Generate and download an Excel template pre-filled with TOWT data for this packing list."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    pl = await _get_pl(token, db, request)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Packing List"

    # Headers
    headers = [
        "CUSTOMER_NAME", "FREIGHT_FORWARDER", "CODE_TRANSITAIRE",
        "SHIPPER_NAME", "PO_NUMBER", "CUSTOMER_BATCH_ID",
        "NOTIFY_ADRESS", "CONSIGNEE_ORDER_ADRESS",
        "PALLET_TYPE", "TYPE_OF_GOODS", "DESCRIPTION_OF_GOODS",
        "BIO_PRODUCTS", "CASES_QUANTITY", "UNITS_PER_CASE",
        "IMO_PRODUCT_CLASS", "PALLET_QUANTITY_PER_BATCH",
        "LENGTH_CM", "WIDTH_CM", "HEIGHT_CM", "WEIGHT_KG",
        "CARGO_VALUE_USD",
    ]

    hdr_fill = PatternFill(start_color="095561", end_color="095561", fill_type="solid")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    yellow_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Pre-fill rows from existing batches
    for i, batch in enumerate(pl.batches, 2):
        row_data = [
            batch.customer_name, batch.freight_forwarder, batch.code_transitaire,
            batch.shipper_name, batch.po_number, batch.customer_batch_id,
            batch.notify_address, batch.consignee_address,
            batch.pallet_type, batch.type_of_goods, batch.description_of_goods,
            batch.bio_products, batch.cases_quantity, batch.units_per_case,
            batch.imo_product_class, batch.pallet_quantity,
            batch.length_cm, batch.width_cm, batch.height_cm, batch.weight_kg,
            batch.cargo_value_usd,
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.fill = yellow_fill
            cell.border = thin_border

    # If no batches, add one empty yellow row
    if not pl.batches:
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=2, column=col, value="")
            cell.fill = yellow_fill
            cell.border = thin_border

    # Auto-width columns
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = max(15, len(headers[col - 1]) + 2)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    ref = pl.order.reference if pl.order else "PL"
    fname = f"PackingList_Template_{ref}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ── Excel Import (external portal) ───────────────────────────
@ext_router.post("/{token}/import", response_class=HTMLResponse)
async def client_import_excel(token: str, request: Request, db: AsyncSession = Depends(get_db)):
    """Import a filled Excel template into the packing list (external portal)."""
    import openpyxl

    pl = await _get_pl(token, db, request)
    if pl.is_locked:
        raise HTTPException(403, detail="Packing list verrouillée")

    form = await request.form()
    file = form.get("file")
    if not file:
        raise HTTPException(400, detail="Aucun fichier")

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(400, detail="Fichier Excel invalide")

    ws = wb.active
    headers = [cell.value for cell in ws[1] if cell.value]
    if not headers:
        raise HTTPException(400, detail="Format invalide — en-têtes manquants")

    header_map = {h: idx for idx, h in enumerate(headers)}

    EXCEL_TO_FIELD = {
        "CUSTOMER_NAME": "customer_name",
        "FREIGHT_FORWARDER": "freight_forwarder",
        "CODE_TRANSITAIRE": "code_transitaire",
        "SHIPPER_NAME": "shipper_name",
        "PO_NUMBER": "po_number",
        "CUSTOMER_BATCH_ID": "customer_batch_id",
        "NOTIFY_ADRESS": "notify_address",
        "CONSIGNEE_ORDER_ADRESS": "consignee_address",
        "PALLET_TYPE": "pallet_type",
        "TYPE_OF_GOODS": "type_of_goods",
        "DESCRIPTION_OF_GOODS": "description_of_goods",
        "BIO_PRODUCTS": "bio_products",
        "CASES_QUANTITY": "cases_quantity",
        "UNITS_PER_CASE": "units_per_case",
        "IMO_PRODUCT_CLASS": "imo_product_class",
        "PALLET_QUANTITY_PER_BATCH": "pallet_quantity",
        "LENGTH_CM": "length_cm",
        "WIDTH_CM": "width_cm",
        "HEIGHT_CM": "height_cm",
        "WEIGHT_KG": "weight_kg",
        "CARGO_VALUE_USD": "cargo_value_usd",
    }

    # Delete existing batches
    for b in pl.batches:
        await db.delete(b)
    await db.flush()

    batch_num = 0
    for row in ws.iter_rows(min_row=2, values_only=False):
        values = [cell.value for cell in row]
        if not any(values):
            continue
        batch_num += 1
        batch = PackingListBatch(packing_list_id=pl.id, batch_number=batch_num)

        # Pre-fill TOWT fields
        batch.booking_confirmation = pl.order.reference
        batch.customer_name = pl.order.client_name
        batch.freight_rate = pl.order.unit_price
        if pl.order.leg:
            leg = pl.order.leg
            batch.voyage_id = leg.leg_code
            batch.vessel = leg.vessel.name if leg.vessel else None
            batch.loading_date = leg.etd.date() if leg.etd else None
            batch.pol_code = leg.departure_port.locode if leg.departure_port else None
            batch.pod_code = leg.arrival_port.locode if leg.arrival_port else None
            batch.pol_name = leg.departure_port.name if leg.departure_port else None
            batch.pod_name = leg.arrival_port.name if leg.arrival_port else None

        # Apply Excel data
        for excel_col, model_field in EXCEL_TO_FIELD.items():
            if excel_col in header_map:
                idx = header_map[excel_col]
                if idx < len(values) and values[idx] is not None:
                    setattr(batch, model_field, values[idx])

        batch.compute_dimensions()
        db.add(batch)

        db.add(PackingListAudit(
            packing_list_id=pl.id, batch_id=None,
            field_name="excel_import",
            old_value=None,
            new_value=f"Batch #{batch_num} importé depuis Excel (portail client)",
            changed_by="Client",
        ))

    await db.flush()
    return RedirectResponse(url=f"/p/{token}/packing?saved=1", status_code=303)
